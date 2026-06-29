from __future__ import annotations

from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from app.utils.permissions import permission_required
from flask import jsonify
from sqlalchemy.orm import joinedload
from flask import Blueprint, flash, redirect, render_template, request, session, url_for
from flask_login import current_user, login_required

from app.extensions import db
from app.models.physical_inventory import PhysicalInventory
from app.models.physical_inventory_line import PhysicalInventoryLine
from app.models.warehouse import Warehouse
from app.models.inventory import WarehouseStock
from app.services.physical_inventory_service import apply_physical_inventory_adjustment
from app.models.article import Article
from io import BytesIO
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


physical_inventory_bp = Blueprint(
    "physical_inventory",
    __name__,
    url_prefix="/physical-inventory",
)


class PhysicalInventoryRouteError(Exception):
    pass


def _to_decimal(value, default="0") -> Decimal:
    try:
        raw = str(value if value is not None else default).strip()
        return Decimal(raw if raw else default)
    except (InvalidOperation, TypeError):
        raise PhysicalInventoryRouteError("Cantidad inválida.")


def _generate_inventory_number() -> str:
    last_id = db.session.query(db.func.max(PhysicalInventory.id)).scalar() or 0
    return f"IF-{int(last_id) + 1:06d}"


@physical_inventory_bp.route("/")
@login_required
@permission_required("inventario_fisico")
def index():
    active_site_id = session.get("active_site_id")

    query = PhysicalInventory.query

    if active_site_id:
        query = query.filter(PhysicalInventory.site_id == int(active_site_id))

    inventories = (
        query
        .order_by(
            PhysicalInventory.created_at.desc(),
            PhysicalInventory.id.desc(),
        )
        .all()
    )

    warehouses_query = Warehouse.query.filter(Warehouse.is_active.is_(True))

    if active_site_id:
        warehouses_query = warehouses_query.filter(Warehouse.site_id == int(active_site_id))

    warehouses = warehouses_query.order_by(Warehouse.name.asc()).all()

    return render_template(
        "physical_inventory/index.html",
        inventories=inventories,
        warehouses=warehouses,
    )


@physical_inventory_bp.route("/create", methods=["POST"])
@login_required
def create():

    active_site_id = session.get("active_site_id")
    warehouse_id = request.form.get("warehouse_id")

    if not active_site_id:
        flash(
            "Debe seleccionar un predio activo antes de crear un inventario físico.",
            "danger",
        )
        return redirect(url_for("physical_inventory.index"))

    if not warehouse_id:
        flash("Debe seleccionar una bodega.", "danger")
        return redirect(url_for("physical_inventory.index"))

    warehouse = Warehouse.query.filter(
        Warehouse.id == int(warehouse_id),
        Warehouse.site_id == int(active_site_id),
        Warehouse.is_active.is_(True),
    ).first()

    if not warehouse:
        flash(
            "La bodega seleccionada no pertenece al predio activo o no está activa.",
            "danger",
        )
        return redirect(url_for("physical_inventory.index"))

    # =====================================================
    # VALIDAR INVENTARIO ABIERTO
    # =====================================================

    existing_open = PhysicalInventory.query.filter(
        PhysicalInventory.warehouse_id == warehouse.id,
        PhysicalInventory.status.in_(["BORRADOR", "EN_CONTEO"]),
    ).first()

    if existing_open:
        flash(
            f"Ya existe un inventario físico abierto para esta bodega: {existing_open.number}.",
            "warning",
        )

        return redirect(
            url_for(
                "physical_inventory.detail",
                inventory_id=existing_open.id,
            )
        )

    # =====================================================
    # VALIDAR STOCK EXISTENTE
    # =====================================================

    stock_exists = (
        db.session.query(WarehouseStock.id)
        .filter(WarehouseStock.warehouse_id == warehouse.id)
        .first()
    )

    if not stock_exists:
        flash(
            "La bodega seleccionada no tiene stock registrado.",
            "warning",
        )
        return redirect(url_for("physical_inventory.index"))

    # =====================================================
    # CREAR CABECERA INVENTARIO
    # =====================================================

    inventory = PhysicalInventory(
        number=_generate_inventory_number(),
        site_id=int(active_site_id),
        warehouse_id=warehouse.id,
        created_by_user_id=current_user.id,
        status="EN_CONTEO",
        notes=request.form.get("notes"),
        created_at=datetime.now(UTC),
    )

    db.session.add(inventory)
    db.session.flush()

    # =====================================================
    # INSERT MASIVO DIRECTO POSTGRESQL
    # =====================================================

    sql = """
    INSERT INTO atm.physical_inventory_lines (
        physical_inventory_id,
        article_id,
        system_quantity,
        physical_quantity,
        difference_quantity
    )
    SELECT
        :inventory_id,
        ws.article_id,
        COALESCE(ws.quantity_on_hand, 0),
        NULL,
        NULL
    FROM atm.warehouse_stock ws
    WHERE ws.warehouse_id = :warehouse_id
    ORDER BY ws.article_id ASC
    """

    db.session.execute(
        db.text(sql),
        {
            "inventory_id": inventory.id,
            "warehouse_id": warehouse.id,
        }
    )

    db.session.commit()

    flash(
        "Inventario físico creado correctamente.",
        "success",
    )

    return redirect(
        url_for(
            "physical_inventory.detail",
            inventory_id=inventory.id,
        )
    )


@physical_inventory_bp.route("/<int:inventory_id>")
@login_required
def detail(inventory_id: int):
    inventory = PhysicalInventory.query.get_or_404(inventory_id)

    active_site_id = session.get("active_site_id")
    if active_site_id and inventory.site_id != int(active_site_id):
        flash("Este inventario físico no pertenece al predio activo.", "danger")
        return redirect(url_for("physical_inventory.index"))

    page = request.args.get("page", 1, type=int)
    per_page = 5000

    pagination = (
        PhysicalInventoryLine.query
        .options(joinedload(PhysicalInventoryLine.article))
        .filter(PhysicalInventoryLine.physical_inventory_id == inventory.id)
        .join(PhysicalInventoryLine.article)
        .order_by(
            db.text("atm.articles.code ASC"),
            PhysicalInventoryLine.id.asc(),
        )
        .paginate(page=page, per_page=per_page, error_out=False)
    )
    lines = pagination.items

    return render_template(
        "physical_inventory/detail.html",
        inventory=inventory,
        lines=lines,
        pagination=pagination,
    )

@physical_inventory_bp.route("/<int:inventory_id>/find-line")
@login_required
def find_line(inventory_id: int):
    inventory = PhysicalInventory.query.get_or_404(inventory_id)

    code = (request.args.get("code") or "").strip().upper()

    if not code:
        return jsonify({"ok": False, "message": "Código vacío."}), 400

    per_page = 5000

    query = (
        PhysicalInventoryLine.query
        .filter(PhysicalInventoryLine.physical_inventory_id == inventory.id)
        .join(PhysicalInventoryLine.article)
        .order_by(
            db.text("atm.articles.code ASC"),
            PhysicalInventoryLine.id.asc(),
        )
    )

    all_ids = [
        row.id
        for row in query.with_entities(PhysicalInventoryLine.id).all()
    ]

    target_line = (
        query
        .filter(
            db.or_(
                db.text("UPPER(atm.articles.code) = :code"),
                db.text("UPPER(atm.articles.barcode) = :code"),
            )
        )
        .params(code=code)
        .first()
    )

    if not target_line:
        return jsonify({
            "ok": False,
            "message": "Artículo no encontrado.",
        })

    try:
        position = all_ids.index(target_line.id)
    except ValueError:
        return jsonify({
            "ok": False,
            "message": "No se pudo ubicar la línea.",
        })

    page = (position // per_page) + 1

    return jsonify({
        "ok": True,
        "page": page,
        "line_id": target_line.id,
        "code": code,
    })

@physical_inventory_bp.route("/update-line", methods=["POST"])
@login_required
def update_line():
    data = request.get_json(silent=True) or {}

    line_id = data.get("line_id")
    field = data.get("field")
    value_raw = data.get("value")

    if not line_id:
        return {"ok": False, "message": "No se recibió la línea."}, 400

    allowed_fields = {
        "count_1_quantity",
        "count_2_quantity",
    }

    if field not in allowed_fields:
        return {"ok": False, "message": "Campo de conteo inválido."}, 400

    line = PhysicalInventoryLine.query.get_or_404(int(line_id))
    inventory = line.physical_inventory

    if inventory.status not in ["BORRADOR", "EN_CONTEO"]:
        return {
            "ok": False,
            "message": "Este inventario físico ya no permite modificaciones.",
        }, 400

    try:
        count_value = _to_decimal(value_raw)
    except PhysicalInventoryRouteError as exc:
        return {"ok": False, "message": str(exc)}, 400

    if count_value < 0:
        return {"ok": False, "message": "La cantidad no puede ser negativa."}, 400

    setattr(line, field, count_value)

    count_1 = line.count_1_quantity or Decimal("0")
    count_2 = line.count_2_quantity or Decimal("0")

    physical_quantity = count_1 + count_2

    system_quantity = line.system_quantity or Decimal("0")
    difference_quantity = physical_quantity - system_quantity

    line.physical_quantity = physical_quantity
    line.difference_quantity = difference_quantity
    line.counted_at = datetime.now(UTC)

    db.session.commit()

    return {
        "ok": True,
        "line_id": line.id,
        "count_1_quantity": str(line.count_1_quantity or Decimal("0")),
        "count_2_quantity": str(line.count_2_quantity or Decimal("0")),
        "system_quantity": str(system_quantity),
        "physical_quantity": str(physical_quantity),
        "difference_quantity": str(difference_quantity),
    }

@physical_inventory_bp.route("/<int:inventory_id>/changes")
@login_required
def get_changes(inventory_id: int):

    inventory = PhysicalInventory.query.get_or_404(inventory_id)

    since = request.args.get("since")

    query = (
        PhysicalInventoryLine.query
        .filter(
            PhysicalInventoryLine.physical_inventory_id == inventory.id,
            PhysicalInventoryLine.counted_at.isnot(None),
        )
    )

    if since:
        try:
            since_dt = datetime.fromisoformat(since)
            query = query.filter(
                PhysicalInventoryLine.counted_at > since_dt
            )
        except Exception:
            pass

    lines = (
        query
        .order_by(PhysicalInventoryLine.counted_at.asc())
        .all()
    )

    return jsonify({
        "ok": True,
        "server_time": datetime.now(UTC).isoformat(),
        "lines": [
            {
                "line_id": line.id,
                "count_1_quantity": (
                    str(line.count_1_quantity)
                    if line.count_1_quantity is not None
                    else ""
                ),
                "count_2_quantity": (
                    str(line.count_2_quantity)
                    if line.count_2_quantity is not None
                    else ""
                ),
                "physical_quantity": (
                    str(line.physical_quantity)
                    if line.physical_quantity is not None
                    else ""
                ),
                "difference_quantity": (
                    str(line.difference_quantity)
                    if line.difference_quantity is not None
                    else ""
                ),
                "counted_at": (
                    line.counted_at.isoformat()
                    if line.counted_at
                    else None
                ),
            }
            for line in lines
        ],
    })

@physical_inventory_bp.route("/<int:inventory_id>/close", methods=["POST"])
@login_required
def close(inventory_id: int):
    inventory = PhysicalInventory.query.get_or_404(inventory_id)

    if inventory.status not in ["BORRADOR", "EN_CONTEO"]:
        flash("Este inventario físico no se puede cerrar en su estado actual.", "warning")
        return redirect(url_for("physical_inventory.detail", inventory_id=inventory.id))

    total_lines = PhysicalInventoryLine.query.filter(
        PhysicalInventoryLine.physical_inventory_id == inventory.id
    ).count()

    counted_lines = PhysicalInventoryLine.query.filter(
        PhysicalInventoryLine.physical_inventory_id == inventory.id,
        PhysicalInventoryLine.physical_quantity.isnot(None),
    ).count()

    if total_lines != counted_lines:
        flash(
            f"No se puede cerrar. Faltan líneas por contar: {total_lines - counted_lines}.",
            "danger",
        )
        return redirect(url_for("physical_inventory.detail", inventory_id=inventory.id))

    inventory.status = "FINALIZADO"
    db.session.commit()

    flash("Inventario físico cerrado correctamente.", "success")
    return redirect(url_for("physical_inventory.detail", inventory_id=inventory.id))


@physical_inventory_bp.route("/<int:inventory_id>/apply-adjustment", methods=["POST"])
@login_required
def apply_adjustment(inventory_id: int):
    inventory = PhysicalInventory.query.get_or_404(inventory_id)

    # 🔒 Validación de estado
    if inventory.status != "FINALIZADO":
        flash("El inventario debe estar FINALIZADO antes de aplicar ajustes.", "danger")
        return redirect(url_for("physical_inventory.detail", inventory_id=inventory.id))

    # 🔒 Confirmación fuerte
    confirmation = request.form.get("confirmation_text", "").strip().upper()

    if confirmation != "AJUSTAR INVENTARIO":
        flash(
            "Debe escribir exactamente 'AJUSTAR INVENTARIO' para confirmar.",
            "danger",
        )
        return redirect(url_for("physical_inventory.detail", inventory_id=inventory.id))

    try:
        apply_physical_inventory_adjustment(
            inventory_id=inventory.id,
            user_id=current_user.id,
        )

        flash("Inventario ajustado correctamente.", "success")

    except Exception as e:
        flash(f"Error al aplicar ajuste: {str(e)}", "danger")

    return redirect(url_for("physical_inventory.detail", inventory_id=inventory.id))

@physical_inventory_bp.route("/<int:inventory_id>/print-report")
@login_required
def print_report(inventory_id):
    inventory = PhysicalInventory.query.get_or_404(inventory_id)

    rows = (
        db.session.query(
            PhysicalInventoryLine,
            Article,
            WarehouseStock.avg_unit_cost,
            WarehouseStock.last_unit_cost,
        )
        .join(Article, Article.id == PhysicalInventoryLine.article_id)
        .outerjoin(
            WarehouseStock,
            db.and_(
                WarehouseStock.article_id == PhysicalInventoryLine.article_id,
                WarehouseStock.warehouse_id == inventory.warehouse_id,
            ),
        )
        .filter(
            PhysicalInventoryLine.physical_inventory_id == inventory.id,
            db.or_(
                PhysicalInventoryLine.count_1_quantity.isnot(None),
                PhysicalInventoryLine.count_2_quantity.isnot(None),
                PhysicalInventoryLine.physical_quantity.isnot(None),
            ),
        )
        .order_by(Article.code.asc())
        .all()
    )

    report_lines = []
    total_difference_amount = Decimal("0")

    for line, article, avg_unit_cost, last_unit_cost in rows:
        system_quantity = line.system_quantity or Decimal("0")
        physical_quantity = line.physical_quantity or Decimal("0")

        # No mostrar artículos donde sistema = 0 y conteo total = 0
        if system_quantity == Decimal("0") and physical_quantity == Decimal("0"):
            continue

        unit_cost = avg_unit_cost or last_unit_cost or Decimal("0")
        difference_quantity = line.difference_quantity or Decimal("0")
        difference_amount = difference_quantity * unit_cost

        total_difference_amount += difference_amount

        report_lines.append({
            "line": line,
            "article": article,
            "unit_cost": unit_cost,
            "difference_amount": difference_amount,
        })

    return render_template(
        "physical_inventory/print_report.html",
        inventory=inventory,
        report_lines=report_lines,
        total_difference_amount=total_difference_amount,
    )

@physical_inventory_bp.route("/update-participants", methods=["POST"])
@login_required
def update_participants():

    data = request.get_json() or {}

    inventory = PhysicalInventory.query.get_or_404(
        data["inventory_id"]
    )

    if inventory.status not in ["BORRADOR", "EN_CONTEO"]:
        return {
            "ok": False
        }, 400

    inventory.participants = (
        data.get("participants") or ""
    ).strip()

    db.session.commit()

    return {
        "ok": True
    }

@physical_inventory_bp.route("/<int:inventory_id>/export-excel")
@login_required
def export_excel(inventory_id):
    inventory = PhysicalInventory.query.get_or_404(inventory_id)

    rows = (
        PhysicalInventoryLine.query
        .options(joinedload(PhysicalInventoryLine.article))
        .filter(PhysicalInventoryLine.physical_inventory_id == inventory.id)
        .join(PhysicalInventoryLine.article)
        .order_by(
            db.text("atm.articles.code ASC"),
            PhysicalInventoryLine.id.asc(),
        )
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario físico"

    ws["A1"] = "Inventario físico"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A2"] = "Número"
    ws["B2"] = inventory.number

    ws["A3"] = "Predio"
    ws["B3"] = inventory.site.name if inventory.site else ""

    ws["A4"] = "Bodega"
    ws["B4"] = inventory.warehouse.name if inventory.warehouse else ""

    ws["A5"] = "Estado"
    ws["B5"] = inventory.status

    headers = [
        "Código",
        "Nombre",
        "Cantidad sistema",
        "Conteo 1",
        "Conteo 2",
        "Total físico",
        "Diferencia",
    ]

    start_row = 7

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")

    for row_idx, line in enumerate(rows, start=start_row + 1):
        ws.cell(row=row_idx, column=1, value=line.article.code if line.article else "")
        ws.cell(row=row_idx, column=2, value=line.article.name if line.article else "")
        ws.cell(row=row_idx, column=3, value=float(line.system_quantity or 0))
        ws.cell(row=row_idx, column=4, value=float(line.count_1_quantity) if line.count_1_quantity is not None else None)
        ws.cell(row=row_idx, column=5, value=float(line.count_2_quantity) if line.count_2_quantity is not None else None)
        ws.cell(row=row_idx, column=6, value=float(line.physical_quantity) if line.physical_quantity is not None else None)
        ws.cell(row=row_idx, column=7, value=float(line.difference_quantity) if line.difference_quantity is not None else None)

    thin = Side(style="thin", color="D9D9D9")

    for row in ws.iter_rows(
        min_row=start_row,
        max_row=start_row + len(rows),
        min_col=1,
        max_col=len(headers),
    ):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center")

    widths = {
        1: 14,
        2: 45,
        3: 16,
        4: 12,
        5: 12,
        6: 14,
        7: 14,
    }

    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A8"
    ws.auto_filter.ref = f"A{start_row}:G{start_row + len(rows)}"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"inventario_fisico_{inventory.number}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )