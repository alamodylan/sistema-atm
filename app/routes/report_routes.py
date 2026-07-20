from datetime import datetime, time

from flask import Blueprint, flash, render_template, request
from flask_login import login_required

from app.extensions import db
from app.models.inventory import InventoryLedger
from app.models.work_order import WorkOrder
from app.models.waste_act import WasteAct
from app.models.article import Article
from app.models.warehouse import Warehouse
from io import BytesIO
from sqlalchemy import func
from flask import send_file
from sqlalchemy.orm import selectinload
from app.models.inventory_entry import InventoryEntry
from app.models.supplier import Supplier


report_bp = Blueprint("reports", __name__)

# =========================================================
# HELPER - CONSULTA MOVIMIENTOS DE INVENTARIO
# =========================================================
def _build_inventory_movements_query(
    date_from="",
    date_to="",
    article_code="",
    warehouse_id="",
    movement_direction="",
):
    query = (
        InventoryLedger.query
        .join(
            Article,
            Article.id == InventoryLedger.article_id,
        )
        .join(
            Warehouse,
            Warehouse.id == InventoryLedger.warehouse_id,
        )
    )

    if date_from:
        parsed_date_from = datetime.combine(
            datetime.strptime(date_from, "%Y-%m-%d").date(),
            time.min,
        )

        query = query.filter(
            InventoryLedger.created_at >= parsed_date_from
        )

    if date_to:
        parsed_date_to = datetime.combine(
            datetime.strptime(date_to, "%Y-%m-%d").date(),
            time.max,
        )

        query = query.filter(
            InventoryLedger.created_at <= parsed_date_to
        )

    if article_code:
        query = query.filter(
            Article.code.ilike(f"%{article_code}%")
        )

    if warehouse_id:
        query = query.filter(
            InventoryLedger.warehouse_id == int(warehouse_id)
        )

    if movement_direction == "IN":
        query = query.filter(
            InventoryLedger.quantity_change > 0
        )

    elif movement_direction == "OUT":
        query = query.filter(
            InventoryLedger.quantity_change < 0
        )

    return query.order_by(
        InventoryLedger.created_at.desc()
    )

# =========================================================
# HOME REPORTES
# =========================================================
@report_bp.route("/", methods=["GET"])
@login_required
def reports_home():
    return render_template(
        "reports/index.html",
        title="Reportes",
        subtitle="Acceda a los reportes principales del sistema.",
    )


# =========================================================
# MOVIMIENTOS DE INVENTARIO
# =========================================================
@report_bp.route("/inventory-movements", methods=["GET"])
@login_required
def inventory_movements_report():
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()
    article_code = (request.args.get("article_code") or "").strip()
    warehouse_id = (request.args.get("warehouse_id") or "").strip()
    movement_direction = (request.args.get("movement_direction") or "").strip()
    page = request.args.get("page", 1, type=int)

    try:
        query = _build_inventory_movements_query(
            date_from=date_from,
            date_to=date_to,
            article_code=article_code,
            warehouse_id=warehouse_id,
            movement_direction=movement_direction,
        )

        if date_from:
            parsed_date_from = datetime.combine(
                datetime.strptime(date_from, "%Y-%m-%d").date(),
                time.min,
            )
            query = query.filter(InventoryLedger.created_at >= parsed_date_from)

        if date_to:
            parsed_date_to = datetime.combine(
                datetime.strptime(date_to, "%Y-%m-%d").date(),
                time.max,
            )
            query = query.filter(InventoryLedger.created_at <= parsed_date_to)

        if article_code:
            query = query.filter(Article.code.ilike(f"%{article_code}%"))

        if warehouse_id:
            query = query.filter(InventoryLedger.warehouse_id == int(warehouse_id))

        if movement_direction == "IN":
            query = query.filter(InventoryLedger.quantity_change > 0)

        if movement_direction == "OUT":
            query = query.filter(InventoryLedger.quantity_change < 0)

        pagination = query.paginate(
            page=page,
            per_page=30,
            error_out=False,
        )
        movements = pagination.items

        work_order_ids = [
            movement.reference_id
            for movement in movements
            if movement.reference_id
            and movement.reference_type == "WORK_ORDER"
        ]

        work_orders_map = {}

        if work_order_ids:
            work_orders = (
                WorkOrder.query
                .filter(WorkOrder.id.in_(work_order_ids))
                .all()
            )

            work_orders_map = {
                work_order.id: work_order
                for work_order in work_orders
            }

        inventory_entry_ids = [
            movement.reference_id
            for movement in movements
            if movement.reference_id
            and movement.reference_type == "INVENTORY_ENTRY"
        ]

        inventory_entries_map = {}

        if inventory_entry_ids:
            inventory_entries = (
                db.session.query(
                    InventoryEntry.id.label("entry_id"),
                    InventoryEntry.invoice_number.label("invoice_number"),
                    func.coalesce(
                        Supplier.commercial_name,
                        Supplier.legal_name,
                        Supplier.code,
                        "-"
                    ).label("supplier_name"),
                )
                .outerjoin(Supplier, Supplier.id == InventoryEntry.supplier_id)
                .filter(InventoryEntry.id.in_(inventory_entry_ids))
                .all()
            )

            inventory_entries_map = {
                entry.entry_id: entry
                for entry in inventory_entries
            }

        for movement in movements:
            movement.equipment_code_report = None
            movement.supplier_name_report = None
            movement.invoice_number_report = None

            # Salidas asociadas a órdenes de trabajo
            if (
                movement.reference_id
                and movement.reference_type == "WORK_ORDER"
            ):
                work_order = work_orders_map.get(movement.reference_id)

                if work_order:
                    movement.equipment_code_report = (
                        work_order.equipment_code_snapshot or None
                    )

            # Entradas creadas normalmente desde el sistema
            if (
                movement.reference_id
                and movement.reference_type == "INVENTORY_ENTRY"
            ):
                entry = inventory_entries_map.get(movement.reference_id)

                if entry:
                    movement.invoice_number_report = (
                        entry.invoice_number or None
                    )

                    movement.supplier_name_report = (
                        entry.supplier_name or None
                    )

            # Entradas importadas mediante SQL
            if movement.reference_type == "FACTURA_COMPRA":
                movement.invoice_number_report = (
                    movement.reference_number or None
                )

                notes = movement.notes or ""

                if "Proveedor:" in notes:
                    movement.supplier_name_report = (
                        notes
                        .split("Proveedor:", 1)[1]
                        .split("|", 1)[0]
                        .strip()
                    ) or None

        warehouses = (
            Warehouse.query
            .order_by(Warehouse.name.asc())
            .all()
        )

        return render_template(
            "reports/inventory_movements.html",
            title="Reporte de movimientos de inventario",
            subtitle="Entradas y salidas por rango de fechas, artículo, predio o bodega.",
            movements=movements,
            pagination=pagination,
            warehouses=warehouses,
            date_from=date_from,
            date_to=date_to,
            article_code=article_code,
            warehouse_id=warehouse_id,
            movement_direction=movement_direction,
        )

    except Exception:
        db.session.rollback()

        flash(
            "Error al cargar el reporte de movimientos.",
            "danger",
        )

        return render_template(
            "reports/inventory_movements.html",
            title="Reporte de movimientos de inventario",
            subtitle="Entradas y salidas por rango de fechas, artículo, predio o bodega.",
            movements=[],
            pagination=None,
            warehouses=[],
            date_from=date_from,
            date_to=date_to,
            article_code=article_code,
            warehouse_id=warehouse_id,
            movement_direction=movement_direction,
        )

# =========================================================
# EXPORTAR MOVIMIENTOS DE INVENTARIO A EXCEL
# =========================================================
@report_bp.route("/inventory-movements/export", methods=["GET"])
@login_required
def export_inventory_movements_report():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()
    article_code = (request.args.get("article_code") or "").strip()
    warehouse_id = (request.args.get("warehouse_id") or "").strip()
    movement_direction = (
        request.args.get("movement_direction") or ""
    ).strip()

    try:
        movements = (
            _build_inventory_movements_query(
                date_from=date_from,
                date_to=date_to,
                article_code=article_code,
                warehouse_id=warehouse_id,
                movement_direction=movement_direction,
            )
            .all()
        )

        # -------------------------------------------------
        # OBTENER EQUIPOS DE ÓRDENES DE TRABAJO
        # -------------------------------------------------
        work_order_ids = {
            movement.reference_id
            for movement in movements
            if movement.reference_id
            and movement.reference_type == "WORK_ORDER"
        }

        work_orders_map = {}

        if work_order_ids:
            work_orders = (
                WorkOrder.query
                .filter(WorkOrder.id.in_(work_order_ids))
                .all()
            )

            work_orders_map = {
                work_order.id: work_order
                for work_order in work_orders
            }

        # -------------------------------------------------
        # OBTENER PROVEEDOR Y FACTURA DE ENTRADAS NORMALES
        # -------------------------------------------------
        inventory_entry_ids = {
            movement.reference_id
            for movement in movements
            if movement.reference_id
            and movement.reference_type == "INVENTORY_ENTRY"
        }

        inventory_entries_map = {}

        if inventory_entry_ids:
            inventory_entries = (
                db.session.query(
                    InventoryEntry.id.label("entry_id"),
                    InventoryEntry.invoice_number.label(
                        "invoice_number"
                    ),
                    func.coalesce(
                        Supplier.commercial_name,
                        Supplier.legal_name,
                        Supplier.code,
                        "-",
                    ).label("supplier_name"),
                )
                .outerjoin(
                    Supplier,
                    Supplier.id == InventoryEntry.supplier_id,
                )
                .filter(
                    InventoryEntry.id.in_(inventory_entry_ids)
                )
                .all()
            )

            inventory_entries_map = {
                entry.entry_id: entry
                for entry in inventory_entries
            }

        # -------------------------------------------------
        # CREAR LIBRO EXCEL
        # -------------------------------------------------
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Entradas y salidas"

        # Título
        worksheet.merge_cells("A1:M1")
        worksheet["A1"] = "Reporte de movimientos de inventario"
        worksheet["A1"].font = Font(
            bold=True,
            size=15,
        )
        worksheet["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        worksheet.row_dimensions[1].height = 25

        # Información de filtros
        worksheet["A3"] = "Fecha de generación"
        worksheet["B3"] = datetime.now().strftime(
            "%d/%m/%Y %H:%M"
        )

        worksheet["A4"] = "Fecha inicial"
        worksheet["B4"] = date_from or "Todas"

        worksheet["A5"] = "Fecha final"
        worksheet["B5"] = date_to or "Todas"

        worksheet["D3"] = "Código artículo"
        worksheet["E3"] = article_code or "Todos"

        worksheet["D4"] = "Bodega"

        warehouse_name = "Todas"

        if warehouse_id:
            warehouse = db.session.get(
                Warehouse,
                int(warehouse_id),
            )

            if warehouse:
                warehouse_name = warehouse.name

        worksheet["E4"] = warehouse_name

        worksheet["D5"] = "Tipo"

        movement_direction_name = {
            "IN": "Solo entradas",
            "OUT": "Solo salidas",
        }.get(
            movement_direction,
            "Entradas y salidas",
        )

        worksheet["E5"] = movement_direction_name

        for cell_reference in [
            "A3",
            "A4",
            "A5",
            "D3",
            "D4",
            "D5",
        ]:
            worksheet[cell_reference].font = Font(bold=True)

        # Encabezados
        header_row = 7

        headers = [
            "Fecha",
            "Movimiento",
            "Código",
            "Artículo",
            "Bodega",
            "Proveedor",
            "Factura",
            "Equipo",
            "Cantidad",
            "Costo unitario",
            "Costo total",
            "Referencia",
            "Usuario",
        ]

        for column_number, header in enumerate(
            headers,
            start=1,
        ):
            cell = worksheet.cell(
                row=header_row,
                column=column_number,
                value=header,
            )

            cell.font = Font(
                bold=True,
                color="FFFFFF",
            )

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="1F4E78",
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        # -------------------------------------------------
        # AGREGAR MOVIMIENTOS
        # -------------------------------------------------
        current_row = header_row + 1

        for movement in movements:
            supplier_name = "-"
            invoice_number = "-"
            equipment_code = "-"

            # Entradas realizadas normalmente desde el sistema
            if (
                movement.reference_id
                and movement.reference_type == "INVENTORY_ENTRY"
            ):
                inventory_entry = inventory_entries_map.get(
                    movement.reference_id
                )

                if inventory_entry:
                    supplier_name = (
                        inventory_entry.supplier_name or "-"
                    )

                    invoice_number = (
                        inventory_entry.invoice_number or "-"
                    )

            # Entradas importadas directamente desde factura
            elif movement.reference_type == "FACTURA_COMPRA":
                if movement.reference_number:
                    invoice_number = movement.reference_number

                notes = movement.notes or ""

                if "Proveedor:" in notes:
                    supplier_name = (
                        notes
                        .split("Proveedor:", 1)[1]
                        .split("|", 1)[0]
                        .strip()
                    ) or "-"

            # Salidas relacionadas con una OT
            if (
                movement.reference_id
                and movement.reference_type == "WORK_ORDER"
            ):
                work_order = work_orders_map.get(
                    movement.reference_id
                )

                if work_order:
                    equipment_code = (
                        work_order.equipment_code_snapshot or "-"
                    )

            reference = (
                movement.reference_number
                or movement.reference_type
                or "-"
            )

            user_name = "-"

            if movement.performed_by_user:
                user_name = (
                    movement.performed_by_user.full_name or "-"
                )

            quantity = float(
                movement.quantity_change or 0
            )

            unit_cost = (
                float(movement.unit_cost)
                if movement.unit_cost is not None
                else None
            )

            total_cost = (
                float(movement.total_cost)
                if movement.total_cost is not None
                else None
            )

            worksheet.append([
                (
                    movement.created_at.strftime(
                        "%d/%m/%Y %H:%M"
                    )
                    if movement.created_at
                    else ""
                ),
                movement.movement_type or "",
                (
                    movement.article.code
                    if movement.article
                    else ""
                ),
                (
                    movement.article.name
                    if movement.article
                    else ""
                ),
                (
                    movement.warehouse.name
                    if movement.warehouse
                    else ""
                ),
                supplier_name,
                invoice_number,
                equipment_code,
                quantity,
                unit_cost,
                total_cost,
                reference,
                user_name,
            ])

            worksheet.cell(
                row=current_row,
                column=9,
            ).number_format = "#,##0.00"

            worksheet.cell(
                row=current_row,
                column=10,
            ).number_format = '₡#,##0.00'

            worksheet.cell(
                row=current_row,
                column=11,
            ).number_format = '₡#,##0.00'

            current_row += 1

        # -------------------------------------------------
        # CONFIGURACIÓN VISUAL
        # -------------------------------------------------
        worksheet.freeze_panes = "A8"
        worksheet.auto_filter.ref = (
            f"A{header_row}:M{max(header_row, current_row - 1)}"
        )

        column_widths = {
            "A": 19,
            "B": 22,
            "C": 15,
            "D": 42,
            "E": 25,
            "F": 30,
            "G": 18,
            "H": 20,
            "I": 13,
            "J": 17,
            "K": 17,
            "L": 22,
            "M": 25,
        }

        for column_letter, width in column_widths.items():
            worksheet.column_dimensions[
                column_letter
            ].width = width

        for row in worksheet.iter_rows(
            min_row=header_row + 1,
            max_row=max(header_row + 1, current_row - 1),
            min_col=1,
            max_col=len(headers),
        ):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename_date = datetime.now().strftime(
            "%Y%m%d_%H%M"
        )

        return send_file(
            output,
            as_attachment=True,
            download_name=(
                f"movimientos_inventario_{filename_date}.xlsx"
            ),
            mimetype=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
        )

    except ValueError:
        db.session.rollback()

        flash(
            "Los filtros utilizados para exportar no son válidos.",
            "warning",
        )

        return inventory_movements_report()

    except Exception as exc:
        db.session.rollback()

        print(
            f"[EXPORT INVENTORY MOVEMENTS ERROR] {exc}"
        )

        flash(
            "Error al exportar los movimientos de inventario.",
            "danger",
        )

        return inventory_movements_report()

# =========================================================
# REPORTE DE ÓRDENES DE TRABAJO
# =========================================================
@report_bp.route("/work-orders", methods=["GET"])
@login_required
def work_orders_report():
    status = (request.args.get("status") or "").strip()
    page = request.args.get("page", 1, type=int)

    try:
        query = WorkOrder.query

        if status:
            query = query.filter(
                WorkOrder.status == status
            )

        pagination = (
            query
            .order_by(WorkOrder.created_at.desc())
            .paginate(
                page=page,
                per_page=15,
                error_out=False,
            )
        )

        return render_template(
            "reports/work_orders.html",
            title="Reporte de órdenes de trabajo",
            subtitle="Historial, estados, tiempos, responsables y mecánicos.",
            work_orders=pagination.items,
            pagination=pagination,
            status=status,
        )

    except Exception as exc:
        print(f"[WORK ORDERS REPORT ERROR] {exc}")

        db.session.rollback()

        flash("Error al cargar el reporte de OT.", "danger")

        return render_template(
            "reports/work_orders.html",
            work_orders=[],
            pagination=None,
            status=status,
        )


# =========================================================
# REPORTE DE ACTAS DE DESECHO
# =========================================================
@report_bp.route("/waste-acts", methods=["GET"])
@login_required
def waste_acts_report():
    status = (request.args.get("status") or "").strip()

    try:
        query = WasteAct.query

        if status:
            query = query.filter(WasteAct.status == status)

        waste_acts = (
            query
            .order_by(WasteAct.created_at.desc())
            .limit(200)
            .all()
        )

        return render_template(
            "reports/waste_acts.html",
            title="Reporte de actas de desecho",
            subtitle="Consulte actas generadas, estados y artículos incluidos.",
            waste_acts=waste_acts,
            status=status,
        )

    except Exception:
        db.session.rollback()

        flash("Error al cargar el reporte de actas.", "danger")

        return render_template(
            "reports/waste_acts.html",
            waste_acts=[],
            status=status,
        )
    

# =========================================================
# HELPER - CONSULTA STOCK HISTÓRICO POR FECHA
# =========================================================
def _build_stock_at_date_query(cutoff_datetime, article_code="", warehouse_id=""):
    quantity_expr = func.coalesce(
        func.sum(InventoryLedger.quantity_change),
        0,
    )

    total_cost_expr = func.coalesce(
        func.sum(
            InventoryLedger.quantity_change *
            func.coalesce(InventoryLedger.unit_cost, 0)
        ),
        0,
    )

    unit_cost_expr = total_cost_expr / func.nullif(quantity_expr, 0)

    query = (
        db.session.query(
            Article.code.label("article_code"),
            Article.name.label("article_name"),
            Warehouse.name.label("warehouse_name"),
            quantity_expr.label("quantity"),
            unit_cost_expr.label("unit_cost"),
            total_cost_expr.label("total_cost"),
        )
        .join(Article, Article.id == InventoryLedger.article_id)
        .join(Warehouse, Warehouse.id == InventoryLedger.warehouse_id)
        .filter(InventoryLedger.created_at <= cutoff_datetime)
        .group_by(
            Article.id,
            Article.code,
            Article.name,
            Warehouse.id,
            Warehouse.name,
        )
        .having(quantity_expr != 0)
    )

    if article_code:
        query = query.filter(Article.code.ilike(f"%{article_code}%"))

    if warehouse_id:
        query = query.filter(InventoryLedger.warehouse_id == int(warehouse_id))

    return query.order_by(Article.code.asc(), Warehouse.name.asc())


# =========================================================
# REPORTE STOCK HISTÓRICO POR FECHA
# =========================================================
@report_bp.route("/stock-at-date", methods=["GET"])
@login_required
def stock_at_date_report():
    cutoff_date = (request.args.get("cutoff_date") or "").strip()
    article_code = (request.args.get("article_code") or "").strip()
    warehouse_id = (request.args.get("warehouse_id") or "").strip()
    page = request.args.get("page", 1, type=int)

    rows = []
    pagination = None

    try:
        warehouses = Warehouse.query.order_by(Warehouse.name.asc()).all()

        if cutoff_date:
            cutoff_datetime = datetime.combine(
                datetime.strptime(cutoff_date, "%Y-%m-%d").date(),
                time.max,
            )

            query = _build_stock_at_date_query(
                cutoff_datetime=cutoff_datetime,
                article_code=article_code,
                warehouse_id=warehouse_id,
            )

            pagination = query.paginate(
                page=page,
                per_page=20,
                error_out=False,
            )

            rows = pagination.items

        return render_template(
            "reports/stock_at_date.html",
            title="Stock histórico por fecha",
            subtitle="Consulte el inventario existente a una fecha específica.",
            rows=rows,
            pagination=pagination,
            warehouses=warehouses,
            cutoff_date=cutoff_date,
            article_code=article_code,
            warehouse_id=warehouse_id,
        )

    except Exception:
        db.session.rollback()

        flash("Error al cargar el reporte de stock histórico.", "danger")

        return render_template(
            "reports/stock_at_date.html",
            title="Stock histórico por fecha",
            subtitle="Consulte el inventario existente a una fecha específica.",
            rows=[],
            pagination=None,
            warehouses=[],
            cutoff_date=cutoff_date,
            article_code=article_code,
            warehouse_id=warehouse_id,
        )


# =========================================================
# EXPORTAR STOCK HISTÓRICO POR FECHA
# =========================================================
@report_bp.route("/stock-at-date/export", methods=["GET"])
@login_required
def export_stock_at_date_report():
    from openpyxl import Workbook

    cutoff_date = (request.args.get("cutoff_date") or "").strip()
    article_code = (request.args.get("article_code") or "").strip()
    warehouse_id = (request.args.get("warehouse_id") or "").strip()

    try:
        if not cutoff_date:
            flash("Debe seleccionar una fecha para exportar.", "warning")
            return stock_at_date_report()

        cutoff_datetime = datetime.combine(
            datetime.strptime(cutoff_date, "%Y-%m-%d").date(),
            time.max,
        )

        rows = (
            _build_stock_at_date_query(
                cutoff_datetime=cutoff_datetime,
                article_code=article_code,
                warehouse_id=warehouse_id,
            )
            .all()
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Stock histórico"

        ws.append([
            "Código artículo",
            "Nombre artículo",
            "Bodega",
            "Cantidad",
            "Costo unitario",
            "Costo total",
            "Fecha consultada",
        ])

        for row in rows:
            quantity = float(row.quantity or 0)
            unit_cost = float(row.unit_cost or 0)
            total_cost = float(row.total_cost or 0)

            ws.append([
                row.article_code,
                row.article_name,
                row.warehouse_name,
                quantity,
                unit_cost,
                total_cost,
                cutoff_date,
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=f"stock_historico_{cutoff_date}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception:
        db.session.rollback()

        flash("Error al exportar el reporte de stock histórico.", "danger")
        return stock_at_date_report()
    
# =========================================================
# HELPER - REPORTE DE OT POR EQUIPO
# =========================================================
def _build_equipment_work_orders_query(equipment_code="", date_from="", date_to=""):
    total_cost_subquery = (
        db.session.query(
            InventoryLedger.reference_id.label("work_order_id"),
            func.coalesce(
                func.sum(func.abs(InventoryLedger.total_cost)),
                0,
            ).label("total_cost"),
        )
        .filter(InventoryLedger.reference_id.isnot(None))
        .filter(InventoryLedger.reference_type == "WORK_ORDER")
        .group_by(InventoryLedger.reference_id)
        .subquery()
    )

    query = (
        db.session.query(
            WorkOrder,
            func.coalesce(total_cost_subquery.c.total_cost, 0).label("total_cost"),
        )
        .outerjoin(
            total_cost_subquery,
            total_cost_subquery.c.work_order_id == WorkOrder.id,
        )
        .options(
            selectinload(WorkOrder.warehouse),
            selectinload(WorkOrder.responsible_user),
        )
    )

    if equipment_code:
        query = query.filter(
            WorkOrder.equipment_code_snapshot.ilike(f"%{equipment_code}%")
        )

    if date_from:
        parsed_date_from = datetime.combine(
            datetime.strptime(date_from, "%Y-%m-%d").date(),
            time.min,
        )
        query = query.filter(WorkOrder.created_at >= parsed_date_from)

    if date_to:
        parsed_date_to = datetime.combine(
            datetime.strptime(date_to, "%Y-%m-%d").date(),
            time.max,
        )
        query = query.filter(WorkOrder.created_at <= parsed_date_to)

    return query.order_by(WorkOrder.created_at.desc())


# =========================================================
# REPORTE DE ÓRDENES DE TRABAJO POR EQUIPO
# =========================================================
@report_bp.route("/equipment-work-orders", methods=["GET"])
@login_required
def equipment_work_orders_report():
    equipment_code = (request.args.get("equipment_code") or "").strip()
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()
    page = request.args.get("page", 1, type=int)

    try:
        rows = []
        pagination = None

        if equipment_code or date_from or date_to:
            query = _build_equipment_work_orders_query(
                equipment_code=equipment_code,
                date_from=date_from,
                date_to=date_to,
            )

            pagination = query.paginate(
                page=page,
                per_page=20,
                error_out=False,
            )

            rows = pagination.items

        return render_template(
            "reports/equipment_work_orders.html",
            title="Reporte de órdenes por equipo",
            subtitle="Consulte órdenes de trabajo creadas a un equipo por rango de fechas.",
            rows=rows,
            pagination=pagination,
            equipment_code=equipment_code,
            date_from=date_from,
            date_to=date_to,
        )

    except Exception:
        db.session.rollback()

        flash("Error al cargar el reporte de órdenes por equipo.", "danger")

        return render_template(
            "reports/equipment_work_orders.html",
            title="Reporte de órdenes por equipo",
            subtitle="Consulte órdenes de trabajo creadas a un equipo por rango de fechas.",
            rows=[],
            pagination=None,
            equipment_code=equipment_code,
            date_from=date_from,
            date_to=date_to,
        )


# =========================================================
# EXPORTAR REPORTE DE ÓRDENES POR EQUIPO
# =========================================================
@report_bp.route("/equipment-work-orders/export", methods=["GET"])
@login_required
def export_equipment_work_orders_report():
    from openpyxl import Workbook

    equipment_code = (request.args.get("equipment_code") or "").strip()
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()

    try:
        rows = (
            _build_equipment_work_orders_query(
                equipment_code=equipment_code,
                date_from=date_from,
                date_to=date_to,
            )
            .all()
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "OT por equipo"

        ws.append([
            "Fecha creación",
            "Número OT",
            "Estado",
            "Tipo reparación",
            "Usuario responsable",
            "Costo total OT",
            "Bodega",
            "Equipo",
        ])

        for row in rows:
            work_order = row.WorkOrder

            ws.append([
                work_order.created_at.strftime("%Y-%m-%d %H:%M") if work_order.created_at else "",
                work_order.number or "",
                work_order.status or "",
                "",
                work_order.responsible_user.full_name if work_order.responsible_user else "",
                float(row.total_cost or 0),
                work_order.warehouse.name if work_order.warehouse else "",
                work_order.equipment_code_snapshot or "",
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="reporte_ot_por_equipo.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception:
        db.session.rollback()

        flash("Error al exportar el reporte de órdenes por equipo.", "danger")
        return equipment_work_orders_report()

# =========================================================
# HELPER - REPORTE DE OT POR CONTENEDOR
# =========================================================
def _build_container_work_orders_query(date_from="", date_to=""):
    total_cost_subquery = (
        db.session.query(
            InventoryLedger.reference_id.label("work_order_id"),
            func.coalesce(
                func.sum(func.abs(InventoryLedger.total_cost)),
                0,
            ).label("total_cost"),
        )
        .filter(InventoryLedger.reference_id.isnot(None))
        .filter(InventoryLedger.reference_type == "WORK_ORDER")
        .group_by(InventoryLedger.reference_id)
        .subquery()
    )

    query = (
        db.session.query(
            WorkOrder,
            func.coalesce(total_cost_subquery.c.total_cost, 0).label("total_cost"),
        )
        .outerjoin(
            total_cost_subquery,
            total_cost_subquery.c.work_order_id == WorkOrder.id,
        )
        .options(
            selectinload(WorkOrder.warehouse),
            selectinload(WorkOrder.responsible_user),
        )
        .filter(
            WorkOrder.equipment_code_snapshot.op("~")(
                r"^[A-Z]{4}-[0-9]{6}-[0-9]{1}$"
            )
        )
    )

    if date_from:
        parsed_date_from = datetime.combine(
            datetime.strptime(date_from, "%Y-%m-%d").date(),
            time.min,
        )
        query = query.filter(WorkOrder.created_at >= parsed_date_from)

    if date_to:
        parsed_date_to = datetime.combine(
            datetime.strptime(date_to, "%Y-%m-%d").date(),
            time.max,
        )
        query = query.filter(WorkOrder.created_at <= parsed_date_to)

    return query.order_by(WorkOrder.created_at.desc())


# =========================================================
# REPORTE DE ÓRDENES POR CONTENEDOR
# =========================================================
@report_bp.route("/container-work-orders", methods=["GET"])
@login_required
def container_work_orders_report():
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()
    page = request.args.get("page", 1, type=int)

    try:
        rows = []
        pagination = None

        if date_from or date_to:
            query = _build_container_work_orders_query(
                date_from=date_from,
                date_to=date_to,
            )

            pagination = query.paginate(
                page=page,
                per_page=20,
                error_out=False,
            )

            rows = pagination.items

        return render_template(
            "reports/container_work_orders.html",
            title="Reporte de órdenes por contenedor",
            subtitle="Consulte las OT asociadas a contenedores por rango de fechas.",
            rows=rows,
            pagination=pagination,
            date_from=date_from,
            date_to=date_to,
        )

    except Exception as exc:
        db.session.rollback()
        print(f"[CONTAINER WORK ORDERS REPORT ERROR] {exc}")

        flash("Error al cargar el reporte de órdenes por contenedor.", "danger")

        return render_template(
            "reports/container_work_orders.html",
            title="Reporte de órdenes por contenedor",
            subtitle="Consulte las OT asociadas a contenedores por rango de fechas.",
            rows=[],
            pagination=None,
            date_from=date_from,
            date_to=date_to,
        )