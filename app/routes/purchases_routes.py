# app/ routes/ purchases_routes.py
from __future__ import annotations

from datetime import datetime, UTC
from decimal import Decimal, InvalidOperation
from zoneinfo import ZoneInfo
from app.extensions import db
from app.models.article_supplier import ArticleSupplier
from app.models.purchase_request_line import PurchaseRequestLine
from io import BytesIO
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.models.purchase_order_approval import PurchaseOrderApproval
from app.models.quotation_line import QuotationLine
from app.models.inventory import WarehouseStock
from sqlalchemy import func
from flask import session
from flask import jsonify
from sqlalchemy.orm import joinedload
from app.models.inventory_entry import InventoryEntry
from app.models.warehouse_location import WarehouseLocation
from app.models.purchase_order import PurchaseOrder
from app.models.purchase_order_line import PurchaseOrderLine
from app.models.quotation_category import QuotationCategory
from app.models.article_quotation_category import ArticleQuotationCategory
from decimal import Decimal
from app.utils.quotation_category_excel import (
    build_category_comparison_excel,
    category_comparison_filename,
)

from flask import (
    Blueprint,
    flash,
    redirect,
    render_template,
    request,
    url_for,
    Response,
)
from flask_login import current_user, login_required

from app.models.article import Article
from app.models.item_category import ItemCategory
from app.models.pending_article import PendingArticle
from app.models.purchase_request import PurchaseRequest
from app.models.site import Site
from app.models.supplier import Supplier
from app.models.unit import Unit
from app.models.warehouse import Warehouse
from app.models.purchase_order import PurchaseOrder
from app.models.purchase_order_candidate import PurchaseOrderCandidate
from app.models.quotation_batch import QuotationBatch
from app.models.purchase_order_line import PurchaseOrderLine
from app.services.inventory_entry_service import (
    InventoryEntryLinePayload,
    InventoryEntryServiceError,
    create_inventory_entry,
    get_inventory_entry_or_404,
    list_inventory_entries,
)
from app.services.pending_article_service import (
    PendingArticleServiceError,
    create_pending_article,
    get_pending_article_or_404,
    list_pending_articles,
    resolve_pending_article,
)
from app.services.purchase_order_service import (
    PurchaseOrderServiceError,
    create_purchase_order,
    create_manual_purchase_order,
    get_purchase_order_or_404,
    list_purchase_orders,
    register_purchase_order_approval,
    adjust_approved_purchase_order_line,
)
from app.services.purchase_request_service import (
    PurchaseRequestLinePayload,
    PurchaseRequestServiceError,
    create_purchase_request,
    get_purchase_request_or_404,
    list_purchase_requests,
    submit_purchase_request,
    list_purchase_requests_for_manager_review,
    update_purchase_request_line_by_manager,
    approve_purchase_request_for_quotation,
    add_purchase_request_lines,
)
from app.services.quotation_service import (
    QuotationLinePayload,
    QuotationServiceError,
    clear_quotation_selection,
    create_quotation_batch,
    create_minimal_supplier_for_quotation,
    create_single_line_quotation,
    get_category_comparison_matrix,
    get_comparison_for_purchase_request_line,
    get_last_price_for_supplier,
    get_quotation_batch_or_404,
    list_pending_purchase_order_candidates_by_supplier,
    list_quotation_batches,
    list_quotation_line_groups,
    list_quotation_request_groups,
    select_quotation_for_purchase_order,
    list_suppliers_for_category_comparison,
    save_category_comparison_matrix,
)

purchases_bp = Blueprint("purchases", __name__, template_folder="../templates")

CR_TZ = ZoneInfo("America/Costa_Rica")


def _to_int(value: str | None) -> int | None:
    value = (value or "").strip()
    return int(value) if value else None


def _to_decimal(value: str | None, default: str = "0") -> Decimal:
    raw = (value or "").strip()
    try:
        return Decimal(raw if raw else default)
    except (InvalidOperation, TypeError):
        raise ValueError("Valor decimal inválido.")


def _get_valid_purchase_orders_for_receiving():
    return (
        PurchaseOrder.query.filter(
            PurchaseOrder.approval_status.in_(["APROBADA", "RECIBIDA_PARCIAL"])
        )
        .order_by(PurchaseOrder.created_at.desc())
    )


def _cr_datetime(value, fmt: str = "%d/%m/%Y %H:%M") -> str:
    if not value:
        return "-"

    try:
        if value.tzinfo is None:
            value = value.replace(tzinfo=UTC)
        return value.astimezone(CR_TZ).strftime(fmt)
    except Exception:
        return "-"


@purchases_bp.route("/")
@login_required
def home():
    return render_template("purchases/home.html")


# =========================
# SOLICITUDES
# =========================
@purchases_bp.route("/requests")
@login_required
def list_requests():
    status = request.args.get("status", type=str)
    priority = request.args.get("priority", type=str)
    search = request.args.get("search", type=str)

    return render_template(
        "purchases/requests/index.html",
        purchase_requests=[],
        selected_status=status,
        selected_priority=priority,
        search=search,
    )

@purchases_bp.route("/requests/partial/list")
@login_required
def list_requests_partial():
    status = request.args.get("status", type=str)
    priority = request.args.get("priority", type=str)
    search = request.args.get("search", type=str)

    try:
        purchase_requests = list_purchase_requests(
            status=status,
            priority=priority,
            search=search,
        )

        return render_template(
            "purchases/requests/_list.html",
            purchase_requests=purchase_requests,
        )

    except Exception as exc:
        print(f"[PURCHASE REQUESTS PARTIAL ERROR] {exc}")
        db.session.rollback()
        return render_template(
            "purchases/requests/_list.html",
            purchase_requests=[],
        ), 500


@purchases_bp.route(
    "/requests/create",
    methods=["GET", "POST"],
)
@login_required
def create_request():
    search = (
        request.args.get("search")
        or ""
    ).strip()

    articles = []

    units = (
        Unit.query
        .order_by(Unit.id.asc())
        .all()
    )

    sites = (
        Site.query
        .filter_by(is_active=True)
        .order_by(Site.name.asc())
        .all()
    )

    warehouses = (
        Warehouse.query
        .filter_by(is_active=True)
        .order_by(Warehouse.name.asc())
        .all()
    )

    # Valores usados para reconstruir el formulario.
    submitted_lines = [
        {
            "article_id": "",
            "article_search": "",
            "pending_article_id": "",
            "manual_name": "",
            "quantity": "",
            "unit_id": "",
            "notes": "",
            "is_urgent": False,
        }
    ]

    line_errors: dict[int, list[str]] = {}

    submitted_header = {
        "priority": "NORMAL",
        "site_id": "",
        "warehouse_id": "",
        "notes": "",
    }

    if request.method == "GET":
        return render_template(
            "purchases/requests/create.html",
            articles=articles,
            units=units,
            sites=sites,
            warehouses=warehouses,
            search=search,
            submitted_lines=submitted_lines,
            line_errors=line_errors,
            submitted_header=submitted_header,
        )

    # =====================================================
    # 1. RECIBIR ENCABEZADO
    # =====================================================

    priority = (
        request.form.get("priority")
        or "NORMAL"
    ).strip().upper()

    notes = (
        request.form.get("notes")
        or ""
    ).strip()

    site_id_raw = (
        request.form.get("site_id")
        or ""
    ).strip()

    warehouse_id_raw = (
        request.form.get("warehouse_id")
        or ""
    ).strip()

    submitted_header = {
        "priority": priority,
        "site_id": site_id_raw,
        "warehouse_id": warehouse_id_raw,
        "notes": notes,
    }

    # =====================================================
    # 2. RECIBIR LÍNEAS SIN GUARDAR NADA TODAVÍA
    # =====================================================

    article_ids = request.form.getlist(
        "line_article_id[]"
    )

    article_searches = request.form.getlist(
        "line_article_search[]"
    )

    pending_article_ids = request.form.getlist(
        "line_pending_article_id[]"
    )

    manual_names = request.form.getlist(
        "line_manual_name[]"
    )

    quantities = request.form.getlist(
        "line_quantity[]"
    )

    unit_ids = request.form.getlist(
        "line_unit_id[]"
    )

    line_notes_list = request.form.getlist(
        "line_notes[]"
    )

    urgent_flags = set(
        request.form.getlist(
            "line_is_urgent[]"
        )
    )

    max_len = max(
        [
            len(article_ids),
            len(article_searches),
            len(pending_article_ids),
            len(manual_names),
            len(quantities),
            len(unit_ids),
            len(line_notes_list),
        ],
        default=0,
    )

    submitted_lines = []

    for index in range(max_len):
        article_id_raw = (
            article_ids[index].strip()
            if index < len(article_ids)
            else ""
        )

        article_search_raw = (
            article_searches[index].strip()
            if index < len(article_searches)
            else ""
        )

        pending_article_id_raw = (
            pending_article_ids[index].strip()
            if index < len(pending_article_ids)
            else ""
        )

        manual_name_raw = (
            manual_names[index].strip()
            if index < len(manual_names)
            else ""
        )

        quantity_raw = (
            quantities[index].strip()
            if index < len(quantities)
            else ""
        )

        unit_id_raw = (
            unit_ids[index].strip()
            if index < len(unit_ids)
            else ""
        )

        line_notes = (
            line_notes_list[index].strip()
            if index < len(line_notes_list)
            else ""
        )

        row = {
            "article_id": article_id_raw,
            "article_search": article_search_raw,
            "pending_article_id": pending_article_id_raw,
            "manual_name": manual_name_raw,
            "quantity": quantity_raw,
            "unit_id": unit_id_raw,
            "notes": line_notes,
            "is_urgent": str(index) in urgent_flags,
        }

        # Ignorar solamente filas completamente vacías.
        if not any(
            [
                article_id_raw,
                article_search_raw,
                pending_article_id_raw,
                manual_name_raw,
                quantity_raw,
                unit_id_raw,
                line_notes,
            ]
        ):
            continue

        submitted_lines.append(row)

    if not submitted_lines:
        submitted_lines = [
            {
                "article_id": "",
                "article_search": "",
                "pending_article_id": "",
                "manual_name": "",
                "quantity": "",
                "unit_id": "",
                "notes": "",
                "is_urgent": False,
            }
        ]

        line_errors[0] = [
            "Debe agregar al menos una línea."
        ]

    # =====================================================
    # 3. VALIDAR ENCABEZADO
    # =====================================================

    valid_priorities = {
        "NORMAL",
        "URGENTE",
        "CRITICA",
    }

    header_errors = []

    if priority not in valid_priorities:
        header_errors.append(
            "La prioridad seleccionada no es válida."
        )

    try:
        site_id = (
            int(site_id_raw)
            if site_id_raw
            else None
        )
    except ValueError:
        site_id = None
        header_errors.append(
            "El predio seleccionado no es válido."
        )

    try:
        warehouse_id = (
            int(warehouse_id_raw)
            if warehouse_id_raw
            else None
        )
    except ValueError:
        warehouse_id = None
        header_errors.append(
            "La bodega seleccionada no es válida."
        )

    # =====================================================
    # 4. VALIDAR TODAS LAS LÍNEAS EN MEMORIA
    # =====================================================

    requested_article_ids: set[int] = set()
    requested_pending_ids: set[int] = set()
    requested_unit_ids: set[int] = set()

    for index, row in enumerate(submitted_lines):
        errors = []

        article_id_raw = row["article_id"]
        pending_article_id_raw = row[
            "pending_article_id"
        ]
        manual_name_raw = row["manual_name"]
        quantity_raw = row["quantity"]
        unit_id_raw = row["unit_id"]

        selected_sources = sum(
            [
                bool(article_id_raw),
                bool(pending_article_id_raw),
                bool(manual_name_raw),
            ]
        )

        if selected_sources == 0:
            errors.append(
                "Debe seleccionar un artículo existente "
                "o escribir un artículo manual."
            )

        if selected_sources > 1:
            errors.append(
                "No puede seleccionar un artículo existente "
                "y escribir uno manual al mismo tiempo."
            )

        if article_id_raw:
            try:
                article_id = int(article_id_raw)
                requested_article_ids.add(article_id)
            except ValueError:
                errors.append(
                    "El artículo seleccionado no es válido."
                )

        if pending_article_id_raw:
            try:
                pending_id = int(
                    pending_article_id_raw
                )
                requested_pending_ids.add(
                    pending_id
                )
            except ValueError:
                errors.append(
                    "El artículo pendiente seleccionado no es válido."
                )

        if not quantity_raw:
            errors.append(
                "Debe indicar una cantidad."
            )
        else:
            try:
                quantity = Decimal(quantity_raw)

                if quantity <= 0:
                    errors.append(
                        "La cantidad debe ser mayor que cero."
                    )

            except (
                InvalidOperation,
                TypeError,
                ValueError,
            ):
                errors.append(
                    "La cantidad indicada no es válida."
                )

        if not unit_id_raw:
            errors.append(
                "Debe seleccionar una unidad."
            )
        else:
            try:
                unit_id = int(unit_id_raw)
                requested_unit_ids.add(unit_id)

            except ValueError:
                errors.append(
                    "La unidad seleccionada no es válida."
                )

        if errors:
            line_errors[index] = errors

    # =====================================================
    # 5. VALIDAR IDs EN CONSULTAS AGRUPADAS
    # =====================================================

    existing_article_ids = set()

    if requested_article_ids:
        existing_article_ids = {
            article_id
            for (article_id,) in (
                db.session.query(Article.id)
                .filter(
                    Article.id.in_(
                        requested_article_ids
                    ),
                    Article.is_active.is_(True),
                )
                .all()
            )
        }

    existing_pending_ids = set()

    if requested_pending_ids:
        existing_pending_ids = {
            pending_id
            for (pending_id,) in (
                db.session.query(
                    PendingArticle.id
                )
                .filter(
                    PendingArticle.id.in_(
                        requested_pending_ids
                    )
                )
                .all()
            )
        }

    existing_unit_ids = set()

    if requested_unit_ids:
        existing_unit_ids = {
            unit_id
            for (unit_id,) in (
                db.session.query(Unit.id)
                .filter(
                    Unit.id.in_(
                        requested_unit_ids
                    )
                )
                .all()
            )
        }

    for index, row in enumerate(submitted_lines):
        errors = line_errors.setdefault(
            index,
            [],
        )

        if row["article_id"]:
            try:
                article_id = int(
                    row["article_id"]
                )

                if (
                    article_id
                    not in existing_article_ids
                ):
                    errors.append(
                        "El artículo seleccionado no existe "
                        "o está inactivo."
                    )

            except ValueError:
                pass

        if row["pending_article_id"]:
            try:
                pending_id = int(
                    row["pending_article_id"]
                )

                if (
                    pending_id
                    not in existing_pending_ids
                ):
                    errors.append(
                        "El artículo pendiente seleccionado "
                        "no existe."
                    )

            except ValueError:
                pass

        if row["unit_id"]:
            try:
                unit_id = int(
                    row["unit_id"]
                )

                if unit_id not in existing_unit_ids:
                    errors.append(
                        "La unidad seleccionada no existe."
                    )

            except ValueError:
                pass

        if not errors:
            line_errors.pop(
                index,
                None,
            )

    # =====================================================
    # 6. SI HAY ERRORES, NO GUARDAR NADA
    # =====================================================

    if header_errors or line_errors:
        for message in header_errors:
            flash(
                message,
                "danger",
            )

        flash(
            "La solicitud no fue guardada. "
            "Revise únicamente las líneas marcadas en rojo; "
            "los demás datos se conservaron.",
            "warning",
        )

        return render_template(
            "purchases/requests/create.html",
            articles=articles,
            units=units,
            sites=sites,
            warehouses=warehouses,
            search=search,
            submitted_lines=submitted_lines,
            line_errors=line_errors,
            submitted_header=submitted_header,
        )

    # =====================================================
    # 7. CONSTRUIR PAYLOADS
    #    TODAVÍA SIN COMMIT
    # =====================================================

    lines: list[
        PurchaseRequestLinePayload
    ] = []

    try:
        for index, row in enumerate(
            submitted_lines
        ):
            article_id = (
                int(row["article_id"])
                if row["article_id"]
                else None
            )

            pending_article_id = (
                int(row["pending_article_id"])
                if row["pending_article_id"]
                else None
            )

            if row["manual_name"]:
                pending = create_pending_article(
                    provisional_name=(
                        row["manual_name"]
                    ),
                    description=row["notes"],
                    category_id=None,
                    unit_id=int(
                        row["unit_id"]
                    ),
                    requested_by_user_id=(
                        current_user.id
                    ),
                    commit=False,
                )

                pending_article_id = pending.id
                article_id = None

            lines.append(
                PurchaseRequestLinePayload(
                    article_id=article_id,
                    pending_article_id=(
                        pending_article_id
                    ),
                    quantity_requested=Decimal(
                        row["quantity"]
                    ),
                    unit_id=int(
                        row["unit_id"]
                    ),
                    line_notes=(
                        row["notes"]
                        or None
                    ),
                    is_urgent=bool(
                        row["is_urgent"]
                    ),
                )
            )

        purchase_request = (
            create_purchase_request(
                requested_by_user_id=(
                    current_user.id
                ),
                priority=priority,
                notes=notes,
                site_id=site_id,
                warehouse_id=warehouse_id,
                lines=lines,
            )
        )

    except PendingArticleServiceError as exc:
        db.session.rollback()

        flash(
            str(exc),
            "danger",
        )

        flash(
            "La solicitud no fue guardada. "
            "Los datos ingresados se conservaron.",
            "warning",
        )

        return render_template(
            "purchases/requests/create.html",
            articles=articles,
            units=units,
            sites=sites,
            warehouses=warehouses,
            search=search,
            submitted_lines=submitted_lines,
            line_errors=line_errors,
            submitted_header=submitted_header,
        )

    except PurchaseRequestServiceError as exc:
        db.session.rollback()

        flash(
            str(exc),
            "danger",
        )

        flash(
            "La solicitud no fue guardada. "
            "Los datos ingresados se conservaron.",
            "warning",
        )

        return render_template(
            "purchases/requests/create.html",
            articles=articles,
            units=units,
            sites=sites,
            warehouses=warehouses,
            search=search,
            submitted_lines=submitted_lines,
            line_errors=line_errors,
            submitted_header=submitted_header,
        )

    except Exception as exc:
        db.session.rollback()

        print(
            "[CREATE PURCHASE REQUEST ERROR] "
            f"error={exc}"
        )

        flash(
            "Ocurrió un error inesperado al guardar "
            "la solicitud. Los datos se conservaron.",
            "danger",
        )

        return render_template(
            "purchases/requests/create.html",
            articles=articles,
            units=units,
            sites=sites,
            warehouses=warehouses,
            search=search,
            submitted_lines=submitted_lines,
            line_errors=line_errors,
            submitted_header=submitted_header,
        )

    flash(
        "Solicitud de compra creada correctamente.",
        "success",
    )

    return redirect(
        url_for(
            "purchases.request_detail",
            request_id=purchase_request.id,
        )
    )

@purchases_bp.route(
    "/requests/<int:request_id>/add-lines",
    methods=["POST"],
)
@login_required
def add_request_lines(request_id: int):
    try:
        article_ids = request.form.getlist(
            "line_article_id[]"
        )

        pending_article_ids = request.form.getlist(
            "line_pending_article_id[]"
        )

        quantities = request.form.getlist(
            "line_quantity[]"
        )

        unit_ids = request.form.getlist(
            "line_unit_id[]"
        )

        line_notes_list = request.form.getlist(
            "line_notes[]"
        )

        urgent_flags = set(
            request.form.getlist(
                "line_is_urgent[]"
            )
        )

        total_lines = max(
            [
                len(article_ids),
                len(pending_article_ids),
                len(quantities),
                len(unit_ids),
                len(line_notes_list),
            ],
            default=0,
        )

        lines: list[PurchaseRequestLinePayload] = []

        for index in range(total_lines):
            article_id_raw = (
                article_ids[index].strip()
                if index < len(article_ids)
                else ""
            )

            pending_article_id_raw = (
                pending_article_ids[index].strip()
                if index < len(pending_article_ids)
                else ""
            )

            quantity_raw = (
                quantities[index].strip()
                if index < len(quantities)
                else ""
            )

            unit_id_raw = (
                unit_ids[index].strip()
                if index < len(unit_ids)
                else ""
            )

            line_notes = (
                line_notes_list[index].strip()
                if index < len(line_notes_list)
                else ""
            )

            if not any(
                [
                    article_id_raw,
                    pending_article_id_raw,
                    quantity_raw,
                    unit_id_raw,
                    line_notes,
                ]
            ):
                continue

            if not article_id_raw and not pending_article_id_raw:
                raise PurchaseRequestServiceError(
                    f"Debe seleccionar un artículo en la línea {index + 1}."
                )

            if not quantity_raw:
                raise PurchaseRequestServiceError(
                    f"Debe indicar la cantidad en la línea {index + 1}."
                )

            if not unit_id_raw:
                raise PurchaseRequestServiceError(
                    f"Debe seleccionar la unidad en la línea {index + 1}."
                )

            try:
                quantity = Decimal(quantity_raw)

            except (InvalidOperation, TypeError, ValueError) as exc:
                raise PurchaseRequestServiceError(
                    f"La cantidad de la línea {index + 1} no es válida."
                ) from exc

            lines.append(
                PurchaseRequestLinePayload(
                    article_id=(
                        int(article_id_raw)
                        if article_id_raw
                        else None
                    ),
                    pending_article_id=(
                        int(pending_article_id_raw)
                        if pending_article_id_raw
                        else None
                    ),
                    quantity_requested=quantity,
                    unit_id=int(unit_id_raw),
                    line_notes=line_notes or None,
                    is_urgent=str(index) in urgent_flags,
                )
            )

        add_purchase_request_lines(
            request_id=request_id,
            requested_by_user_id=current_user.id,
            lines=lines,
        )

        flash(
            "Artículo agregado correctamente a la solicitud.",
            "success",
        )

    except PurchaseRequestServiceError as exc:
        db.session.rollback()
        flash(str(exc), "danger")

    except Exception as exc:
        db.session.rollback()

        print(
            "[ADD PURCHASE REQUEST LINES ERROR] "
            f"request_id={request_id} "
            f"error={exc}"
        )

        flash(
            "Ocurrió un error al agregar el artículo a la solicitud.",
            "danger",
        )

    return redirect(
        url_for(
            "purchases.request_detail",
            request_id=request_id,
        )
    )

@purchases_bp.route("/requests/articles/search")
@login_required
def search_request_articles():
    term = (request.args.get("q") or "").strip()
    warehouse_id = _to_int(request.args.get("warehouse_id"))

    if not warehouse_id:
        return {"items": []}

    if len(term) < 2:
        return {"items": []}

    search_like = f"%{term}%"

    try:
        rows = (
            db.session.query(Article, WarehouseStock)
            .outerjoin(
                WarehouseStock,
                db.and_(
                    WarehouseStock.article_id == Article.id,
                    WarehouseStock.warehouse_id == warehouse_id,
                )
            )
            .filter(
                Article.is_active.is_(True),
                db.or_(
                    Article.code.ilike(search_like),
                    Article.name.ilike(search_like),
                )
            )
            .order_by(Article.code.asc())
            .limit(30)
            .all()
        )

        items = []

        for article, stock in rows:
            quantity_on_hand = Decimal(str(stock.quantity_on_hand or 0)) if stock else Decimal("0")
            available_quantity = Decimal(str(stock.available_quantity or 0)) if stock else Decimal("0")

            unit_name = ""
            if article.unit:
                unit_name = (
                    getattr(article.unit, "name", None)
                    or getattr(article.unit, "code", None)
                    or ""
                )

            items.append({
                "id": article.id,
                "code": article.code,
                "name": article.name,
                "unit_id": article.unit_id,
                "unit_name": unit_name,
                "quantity_on_hand": str(quantity_on_hand),
                "available_quantity": str(available_quantity),
            })

        return {"items": items}

    except Exception as exc:
        print(f"[SEARCH REQUEST ARTICLES ERROR] {exc}")
        db.session.rollback()
        return {"items": []}, 500

@purchases_bp.route(
    "/requests/<int:request_id>",
    methods=["GET"],
)
@login_required
def request_detail(request_id: int):
    purchase_request = get_purchase_request_or_404(
        request_id
    )

    units = (
        Unit.query
        .order_by(
            Unit.id.asc()
        )
        .all()
    )

    return render_template(
        "purchases/requests/detail.html",
        purchase_request=purchase_request,
        units=units,
    )
@purchases_bp.route("/requests/<int:request_id>/send", methods=["POST"])
@login_required
def send_request(request_id: int):
    try:
        purchase_request = submit_purchase_request(request_id=request_id)

    except PurchaseRequestServiceError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("purchases.request_detail", request_id=request_id))

    if purchase_request.status == "EN_REVISION_PROVEEDURIA":
        flash("Solicitud enviada directamente a proveeduría correctamente.", "success")
    else:
        flash("Solicitud enviada a jefatura correctamente.", "success")

    return redirect(
        url_for("purchases.request_detail", request_id=purchase_request.id)
    )

@purchases_bp.route("/manager/purchase-requests")
@login_required
def manager_purchase_requests():
    active_site_id = session.get("active_site_id")

    if not active_site_id:
        purchase_requests = []
    else:
        active_site_id = int(active_site_id)

        purchase_requests = (
            PurchaseRequest.query
            .filter(
                db.or_(
                    PurchaseRequest.review_site_id == active_site_id,
                    db.and_(
                        PurchaseRequest.review_site_id.is_(None),
                        PurchaseRequest.site_id == active_site_id,
                        PurchaseRequest.sent_direct_to_procurement.is_(False),
                    ),
                ),
                PurchaseRequest.status == "ENVIADA",
            )
            .order_by(
                PurchaseRequest.created_at.desc(),
                PurchaseRequest.id.desc(),
            )
            .all()
        )

    return render_template(
        "dashboard/manager.html",
        purchase_requests=purchase_requests,
    )

@purchases_bp.route("/dashboard/manager/purchase-request-lines/<int:line_id>/update", methods=["POST"])
@login_required
def manager_update_request_line(line_id: int):
    quantity_raw = request.form.get("quantity_requested")
    cancel_line = request.form.get("cancel_line") == "1"

    line = PurchaseRequestLine.query.get_or_404(line_id)

    try:
        quantity = Decimal(quantity_raw or "0")

        updated_line = update_purchase_request_line_by_manager(
            line_id=line_id,
            quantity_requested=quantity,
            cancel_line=cancel_line,
        )
    except PurchaseRequestServiceError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("dashboard.manager_dashboard"))

    flash("Línea actualizada correctamente.", "success")
    return redirect(url_for("dashboard.manager_dashboard"))


@purchases_bp.route("/dashboard/manager/purchase-requests/<int:request_id>/approve", methods=["POST"])
@login_required
def manager_approve_request(request_id: int):
    line_ids = request.form.getlist("line_id[]")
    quantities = request.form.getlist("quantity_requested[]")
    cancelled_lines = set(request.form.getlist("cancel_line[]"))

    review_lines = []

    for index, raw_line_id in enumerate(line_ids):
        line_id = _to_int(raw_line_id)

        if not line_id:
            continue

        quantity_raw = quantities[index] if index < len(quantities) else None

        review_lines.append(
            {
                "line_id": line_id,
                "quantity_requested": quantity_raw,
                "cancel_line": str(line_id) in cancelled_lines,
            }
        )

    try:
        approve_purchase_request_for_quotation(
            request_id=request_id,
            review_lines=review_lines,
        )
    except PurchaseRequestServiceError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("dashboard.manager_dashboard"))

    flash("Solicitud aprobada y enviada a proveeduría correctamente.", "success")
    return redirect(url_for("dashboard.manager_dashboard"))
# =========================
# PENDING ARTICLES
# =========================
@purchases_bp.route("/pending-articles")
@login_required
def list_pending_articles_route():
    status = request.args.get("status", type=str)
    search = request.args.get("search", type=str)

    pending_articles = list_pending_articles(status=status, search=search)

    return render_template(
        "purchases/pending_articles/index.html",
        pending_articles=pending_articles,
        selected_status=status,
        search=search,
    )


@purchases_bp.route("/pending-articles/create", methods=["GET", "POST"])
@login_required
def create_pending_article_route():
    categories = ItemCategory.query.order_by(ItemCategory.name.asc()).all()
    units = Unit.query.order_by(Unit.id.asc()).all()

    if request.method == "POST":
        try:
            pending_article = create_pending_article(
                provisional_name=request.form.get("provisional_name"),
                description=request.form.get("description"),
                category_id=_to_int(request.form.get("category_id")),
                unit_id=_to_int(request.form.get("unit_id")),
                requested_by_user_id=current_user.id,
            )
        except PendingArticleServiceError as exc:
            flash(str(exc), "danger")
            return render_template(
                "purchases/pending_articles/create.html",
                categories=categories,
                units=units,
            )

        flash("Artículo pendiente creado correctamente.", "success")
        return redirect(url_for("purchases.pending_article_detail", pending_article_id=pending_article.id))

    return render_template(
        "purchases/pending_articles/create.html",
        categories=categories,
        units=units,
    )


@purchases_bp.route("/pending-articles/<int:pending_article_id>")
@login_required
def pending_article_detail(pending_article_id: int):
    pending_article = get_pending_article_or_404(pending_article_id)

    return render_template(
        "purchases/pending_articles/detail.html",
        pending_article=pending_article,
    )


@purchases_bp.route("/pending-articles/<int:pending_article_id>/resolve", methods=["POST"])
@login_required
def resolve_pending_article_route(pending_article_id: int):
    final_code = (request.form.get("final_code") or "").strip()
    final_name = (request.form.get("final_name") or "").strip()

    if not final_code:
        flash("Debes indicar el código definitivo de 5 dígitos.", "danger")
        return redirect(url_for("purchases.pending_article_detail", pending_article_id=pending_article_id))

    if not final_name:
        flash("Debes indicar el nombre definitivo del artículo.", "danger")
        return redirect(url_for("purchases.pending_article_detail", pending_article_id=pending_article_id))

    try:
        resolve_pending_article(
            pending_article_id=pending_article_id,
            final_code=final_code,
            final_name=final_name,
        )
    except PendingArticleServiceError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("purchases.pending_article_detail", pending_article_id=pending_article_id))

    flash("Artículo pendiente resuelto correctamente.", "success")
    return redirect(url_for("purchases.pending_article_detail", pending_article_id=pending_article_id))


# =========================
# COTIZACIONES
# =========================
@purchases_bp.route("/quotations")
@login_required
def list_quotations():
    search = request.args.get("search", type=str)

    quotation_requests = list_quotation_request_groups(search=search)

    return render_template(
        "purchases/quotations/index.html",
        quotation_requests=quotation_requests,
        search=search,
    )

@purchases_bp.route("/quotations/request/<int:request_id>")
@login_required
def quotation_request_lines(request_id: int):
    purchase_request = PurchaseRequest.query.get_or_404(
        request_id
    )

    edit_categories = (
        request.args.get("edit_categories") == "1"
    )

    # =====================================================
    # 1. CARGAR TODAS LAS LÍNEAS
    # =====================================================
    request_lines = (
        PurchaseRequestLine.query
        .options(
            joinedload(PurchaseRequestLine.article),
            joinedload(PurchaseRequestLine.pending_article),
            joinedload(PurchaseRequestLine.unit),
        )
        .filter(
            PurchaseRequestLine.purchase_request_id
            == request_id,
            PurchaseRequestLine.line_status
            != "CANCELADA",
        )
        .order_by(
            PurchaseRequestLine.id.asc()
        )
        .all()
    )

    # El catálogo es pequeño: actualmente solo 15 registros.
    categories = (
        QuotationCategory.query
        .filter(
            QuotationCategory.is_active.is_(True)
        )
        .order_by(
            QuotationCategory.sort_order.asc(),
            QuotationCategory.name.asc(),
            QuotationCategory.id.asc(),
        )
        .all()
    )

    if not request_lines:
        return render_template(
            "purchases/quotations/request_lines.html",
            purchase_request=purchase_request,
            line_groups=[],
            grouped_line_groups=[],
            categories=categories,
            edit_categories=edit_categories,
        )

    line_ids = [
        line.id
        for line in request_lines
    ]

    article_ids = {
        line.article_id
        for line in request_lines
        if line.article_id is not None
    }

    pending_article_ids = {
        line.pending_article_id
        for line in request_lines
        if line.pending_article_id is not None
    }

    # =====================================================
    # 2. CONTAR COTIZACIONES EN UNA SOLA CONSULTA
    # =====================================================
    quotation_rows = (
        db.session.query(
            QuotationLine.purchase_request_line_id.label(
                "purchase_request_line_id"
            ),
            func.count(
                QuotationLine.id
            ).label("total_quotes"),
            func.sum(
                db.case(
                    (
                        QuotationLine.status == "BORRADOR",
                        1,
                    ),
                    else_=0,
                )
            ).label("draft_quotes"),
            func.sum(
                db.case(
                    (
                        QuotationLine.status == "COTIZADA",
                        1,
                    ),
                    else_=0,
                )
            ).label("confirmed_quotes"),
            func.max(
                QuotationLine.quote_date
            ).label("last_quote_date"),
        )
        .filter(
            QuotationLine.purchase_request_line_id.in_(
                line_ids
            )
        )
        .group_by(
            QuotationLine.purchase_request_line_id
        )
        .all()
    )

    quotation_map = {
        row.purchase_request_line_id: row
        for row in quotation_rows
    }

    # =====================================================
    # 3. CARGAR TODAS LAS CATEGORÍAS ASIGNADAS
    #    EN UNA SOLA CONSULTA
    # =====================================================
    assignment_filters = []

    if article_ids:
        assignment_filters.append(
            ArticleQuotationCategory.article_id.in_(
                article_ids
            )
        )

    if pending_article_ids:
        assignment_filters.append(
            ArticleQuotationCategory.pending_article_id.in_(
                pending_article_ids
            )
        )

    article_category_map = {}
    pending_category_map = {}

    if assignment_filters:
        assignment_rows = (
            db.session.query(
                ArticleQuotationCategory.article_id,
                ArticleQuotationCategory.pending_article_id,
                QuotationCategory.id.label(
                    "category_id"
                ),
                QuotationCategory.name.label(
                    "category_name"
                ),
                QuotationCategory.sort_order.label(
                    "category_sort_order"
                ),
            )
            .join(
                QuotationCategory,
                QuotationCategory.id
                == ArticleQuotationCategory.quotation_category_id,
            )
            .filter(
                db.or_(*assignment_filters)
            )
            .all()
        )

        for row in assignment_rows:
            category_data = {
                "id": row.category_id,
                "name": row.category_name,
                "sort_order": row.category_sort_order,
            }

            if row.article_id is not None:
                article_category_map[
                    row.article_id
                ] = category_data

            if row.pending_article_id is not None:
                pending_category_map[
                    row.pending_article_id
                ] = category_data

    # =====================================================
    # 4. CONSTRUIR LAS FILAS SIN CONSULTAS EN EL FOR
    # =====================================================
    line_groups = []

    for request_line in request_lines:
        quotation_data = quotation_map.get(
            request_line.id
        )

        category_data = None
        item_type = None
        item_id = None

        if request_line.article_id is not None:
            item_type = "ARTICLE"
            item_id = request_line.article_id
            category_data = article_category_map.get(
                request_line.article_id
            )

        elif request_line.pending_article_id is not None:
            item_type = "PENDING"
            item_id = request_line.pending_article_id
            category_data = pending_category_map.get(
                request_line.pending_article_id
            )

        line_groups.append(
            {
                "purchase_request_id": (
                    request_line.purchase_request_id
                ),
                "purchase_request_number": (
                    purchase_request.number
                ),
                "purchase_request_line_id": (
                    request_line.id
                ),
                "article_id": (
                    request_line.article_id
                ),
                "pending_article_id": (
                    request_line.pending_article_id
                ),
                "item_type": item_type,
                "item_id": item_id,
                "item_code": (
                    request_line.item_code or "-"
                ),
                "item_name": (
                    request_line.item_name
                    or "Sin artículo"
                ),
                "quantity_requested": (
                    request_line.quantity_requested
                ),
                "line_status": (
                    request_line.line_status
                ),
                "total_quotes": int(
                    quotation_data.total_quotes
                    if quotation_data
                    else 0
                ),
                "draft_quotes": int(
                    quotation_data.draft_quotes
                    if quotation_data
                    and quotation_data.draft_quotes
                    is not None
                    else 0
                ),
                "confirmed_quotes": int(
                    quotation_data.confirmed_quotes
                    if quotation_data
                    and quotation_data.confirmed_quotes
                    is not None
                    else 0
                ),
                "best_price": None,
                "best_supplier": None,
                "last_quote_date": (
                    quotation_data.last_quote_date
                    if quotation_data
                    else None
                ),
                "quotation_category_id": (
                    category_data["id"]
                    if category_data
                    else None
                ),
                "quotation_category_name": (
                    category_data["name"]
                    if category_data
                    else "SIN CATEGORÍA"
                ),
                "quotation_category_sort_order": (
                    category_data["sort_order"]
                    if category_data
                    else 999999
                ),
            }
        )

    # =====================================================
    # 5. ORDENAR Y AGRUPAR EN MEMORIA
    # =====================================================
    line_groups.sort(
        key=lambda item: (
            item[
                "quotation_category_sort_order"
            ],
            item[
                "quotation_category_name"
            ],
            item["item_name"],
            item["purchase_request_line_id"],
        )
    )

    grouped_line_groups = []

    current_group_key = object()
    current_group = None

    for item in line_groups:
        group_key = (
            item["quotation_category_id"]
            if item["quotation_category_id"]
            is not None
            else "SIN_CATEGORIA"
        )

        if group_key != current_group_key:
            current_group = {
                "category_id": (
                    item[
                        "quotation_category_id"
                    ]
                ),
                "category_name": (
                    item[
                        "quotation_category_name"
                    ]
                ),
                "lines": [],
            }

            grouped_line_groups.append(
                current_group
            )

            current_group_key = group_key

        current_group["lines"].append(item)

    return render_template(
        "purchases/quotations/request_lines.html",
        purchase_request=purchase_request,
        line_groups=line_groups,
        grouped_line_groups=grouped_line_groups,
        categories=categories,
        edit_categories=edit_categories,
    )

@purchases_bp.route(
    "/quotations/request/<int:request_id>/category/<int:category_id>"
)
@login_required
def quotation_category_comparison(
    request_id: int,
    category_id: int,
):
    """
    Abre el comparativo completo de una categoría dentro de una
    solicitud de compra.

    category_id:
    - ID real de atm.quotation_categories.
    - El valor 0 representa líneas sin categoría.
    """

    purchase_request = PurchaseRequest.query.get_or_404(
        request_id
    )

    quotation_category = None

    if category_id != 0:
        quotation_category = (
            QuotationCategory.query
            .filter(
                QuotationCategory.id == category_id,
                QuotationCategory.is_active.is_(True),
            )
            .first()
        )

        if not quotation_category:
            flash(
                "La categoría indicada no existe o está inactiva.",
                "danger",
            )

            return redirect(
                url_for(
                    "purchases.quotation_request_lines",
                    request_id=request_id,
                )
            )

    try:
        comparison_matrix = (
            get_category_comparison_matrix(
                purchase_request_id=request_id,
                quotation_category_id=(
                    category_id
                    if category_id != 0
                    else None
                ),
            )
        )

    except QuotationServiceError as exc:
        flash(
            str(exc),
            "danger",
        )

        return redirect(
            url_for(
                "purchases.quotation_request_lines",
                request_id=request_id,
            )
        )

    except Exception as exc:
        db.session.rollback()

        print(
            "[QUOTATION CATEGORY COMPARISON ERROR] "
            f"{exc}"
        )

        flash(
            "No fue posible cargar el comparativo de la categoría.",
            "danger",
        )

        return redirect(
            url_for(
                "purchases.quotation_request_lines",
                request_id=request_id,
            )
        )

    return render_template(
        "purchases/quotations/category_comparison.html",
        purchase_request=purchase_request,
        quotation_category=quotation_category,
        category_id=category_id,
        comparison_matrix=comparison_matrix,
        edit_mode=False,
    )

@purchases_bp.route(
    "/quotations/request/<int:request_id>/category/<int:category_id>/suppliers"
)
@login_required
def quotation_category_suppliers(
    request_id: int,
    category_id: int,
):

    exclude_supplier_ids = []

    for raw_supplier_id in request.args.getlist(
        "exclude_supplier_id"
    ):
        supplier_id = _to_int(raw_supplier_id)

        if supplier_id:
            exclude_supplier_ids.append(
                supplier_id
            )

    search = (
        request.args.get("search")
        or ""
    ).strip()

    try:

        suppliers = (
            list_suppliers_for_category_comparison(
                purchase_request_id=request_id,
                quotation_category_id=(
                    None
                    if category_id == 0
                    else category_id
                ),
                exclude_supplier_ids=exclude_supplier_ids,
                search=search,
            )
        )

        return jsonify(
            {
                "items": suppliers
            }
        )

    except QuotationServiceError as exc:

        return jsonify(
            {
                "items": [],
                "error": str(exc),
            }
        ), 400

    except Exception as exc:

        db.session.rollback()

        print(
            "[CATEGORY SUPPLIERS ERROR]",
            exc,
        )

        return jsonify(
            {
                "items": [],
                "error":
                    "No fue posible cargar los proveedores.",
            }
        ), 500

@purchases_bp.route(
    "/quotations/request/<int:request_id>/category/<int:category_id>/save",
    methods=["POST"],
)
@login_required
def save_category_comparison(
    request_id: int,
    category_id: int,
):

    payload = (
        request.get_json(silent=True)
        or {}
    )

    supplier_columns = payload.get(
        "supplier_columns",
        [],
    )

    try:

        result = (
            save_category_comparison_matrix(
                purchase_request_id=request_id,
                quotation_category_id=(
                    None
                    if category_id == 0
                    else category_id
                ),
                created_by_user_id=current_user.id,
                supplier_columns=supplier_columns,
            )
        )

        return jsonify(
            {
                "success": True,
                **result,
            }
        )

    except QuotationServiceError as exc:

        return jsonify(
            {
                "success": False,
                "error": str(exc),
            }
        ), 400

    except Exception as exc:

        db.session.rollback()

        print(
            "[SAVE CATEGORY COMPARISON ERROR]",
            exc,
        )

        return jsonify(
            {
                "success": False,
                "error":
                    "No fue posible guardar el comparativo.",
            }
        ), 500

@purchases_bp.route(
    "/quotations/request/<int:request_id>/category/<int:category_id>/excel"
)
@login_required
def export_category_comparison_excel(
    request_id: int,
    category_id: int,
):
    """
    Exporta el comparativo horizontal de una categoría específica.

    El valor 0 representa las líneas sin categoría.
    """
    purchase_request = PurchaseRequest.query.get_or_404(
        request_id
    )

    try:
        comparison_matrix = get_category_comparison_matrix(
            purchase_request_id=request_id,
            quotation_category_id=(
                None
                if category_id == 0
                else category_id
            ),
        )

        output = build_category_comparison_excel(
            comparison_matrix=comparison_matrix,
        )

        download_name = category_comparison_filename(
            purchase_request_number=purchase_request.number,
            category_label=comparison_matrix.get(
                "category_label"
            ),
        )

        return send_file(
            output,
            as_attachment=True,
            download_name=download_name,
            mimetype=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

    except QuotationServiceError as exc:
        flash(str(exc), "danger")

    except Exception as exc:
        db.session.rollback()
        print(
            "[EXPORT CATEGORY COMPARISON EXCEL ERROR] "
            f"{exc}"
        )
        flash(
            "No fue posible exportar el comparativo a Excel.",
            "danger",
        )

    return redirect(
        url_for(
            "purchases.quotation_category_comparison",
            request_id=request_id,
            category_id=category_id,
        )
    )

@purchases_bp.route(
    "/quotations/request/<int:request_id>/category/<int:category_id>/print"
)
@login_required
def print_category_comparison(
    request_id: int,
    category_id: int,
):
    """
    Imprime el comparativo de una categoría específica.
    """

    purchase_request = PurchaseRequest.query.get_or_404(request_id)

    comparison_matrix = get_category_comparison_matrix(
        purchase_request_id=request_id,
        quotation_category_id=category_id,
    )

    return render_template(
        "purchases/quotations/print_category_comparison.html",
        purchase_request=purchase_request,
        comparison_matrix=comparison_matrix,
        category_id=category_id,
        printed_at=datetime.now(),
    )

@purchases_bp.route(
    "/quotations/request-line/<int:line_id>/select",
    methods=["POST"],
)
@login_required
def select_quotation_candidate(line_id: int):
    """
    Selecciona manualmente una cotización como candidata para
    generar una orden de compra.

    La selección no crea todavía la OC. Únicamente crea o actualiza
    el PurchaseOrderCandidate correspondiente a la línea.
    """

    quotation_line_id = _to_int(
        request.form.get("quotation_line_id")
    )

    request_id = _to_int(
        request.form.get("request_id")
    )

    category_id = _to_int(
        request.form.get("category_id")
    )

    if not quotation_line_id:
        flash(
            "Debe seleccionar una cotización válida.",
            "danger",
        )

        if request_id is not None:
            return redirect(
                url_for(
                    "purchases.quotation_category_comparison",
                    request_id=request_id,
                    category_id=(
                        category_id
                        if category_id is not None
                        else 0
                    ),
                )
            )

        return redirect(
            url_for(
                "purchases.quotation_line_view",
                line_id=line_id,
            )
        )

    try:
        select_quotation_for_purchase_order(
            purchase_request_line_id=line_id,
            quotation_line_id=quotation_line_id,
            selected_by_user_id=current_user.id,
        )

    except QuotationServiceError as exc:
        flash(
            str(exc),
            "danger",
        )

    else:
        flash(
            "Cotización seleccionada correctamente para generar la orden de compra.",
            "success",
        )

    if request_id is not None:
        return redirect(
            url_for(
                "purchases.quotation_category_comparison",
                request_id=request_id,
                category_id=(
                    category_id
                    if category_id is not None
                    else 0
                ),
            )
        )

    return redirect(
        url_for(
            "purchases.quotation_line_view",
            line_id=line_id,
        )
    )


@purchases_bp.route(
    "/quotations/request-line/<int:line_id>/selection/clear",
    methods=["POST"],
)
@login_required
def clear_quotation_candidate(line_id: int):
    """
    Cancela una selección de pre-OC mientras todavía no haya sido
    utilizada para crear una orden de compra.
    """

    request_id = _to_int(
        request.form.get("request_id")
    )

    category_id = _to_int(
        request.form.get("category_id")
    )

    try:
        clear_quotation_selection(
            purchase_request_line_id=line_id,
            cancelled_by_user_id=current_user.id,
        )

    except QuotationServiceError as exc:
        flash(
            str(exc),
            "danger",
        )

    else:
        flash(
            "La selección de cotización fue eliminada correctamente.",
            "success",
        )

    if request_id is not None:
        return redirect(
            url_for(
                "purchases.quotation_category_comparison",
                request_id=request_id,
                category_id=(
                    category_id
                    if category_id is not None
                    else 0
                ),
            )
        )

    return redirect(
        url_for(
            "purchases.quotation_line_view",
            line_id=line_id,
        )
    )

@purchases_bp.route(
    "/quotations/request/<int:request_id>/categories/save",
    methods=["POST"],
)
@login_required
def save_quotation_request_categories(
    request_id: int,
):
    PurchaseRequest.query.get_or_404(
        request_id
    )

    item_types = request.form.getlist(
        "item_type[]"
    )

    item_ids = request.form.getlist(
        "item_id[]"
    )

    category_ids = request.form.getlist(
        "quotation_category_id[]"
    )

    if not (
        len(item_types)
        == len(item_ids)
        == len(category_ids)
    ):
        flash(
            "La información de categorías está incompleta.",
            "danger",
        )

        return redirect(
            url_for(
                "purchases.quotation_request_lines",
                request_id=request_id,
                edit_categories=1,
            )
        )

    # =====================================================
    # 1. VALIDAR LOS ARTÍCULOS QUE PERTENECEN A LA SOLICITUD
    # =====================================================
    request_item_rows = (
        db.session.query(
            PurchaseRequestLine.article_id,
            PurchaseRequestLine.pending_article_id,
        )
        .filter(
            PurchaseRequestLine.purchase_request_id
            == request_id,
            PurchaseRequestLine.line_status
            != "CANCELADA",
        )
        .all()
    )

    valid_article_ids = {
        row.article_id
        for row in request_item_rows
        if row.article_id is not None
    }

    valid_pending_article_ids = {
        row.pending_article_id
        for row in request_item_rows
        if row.pending_article_id is not None
    }

    submitted_rows = []
    submitted_category_ids = set()

    for index, raw_item_type in enumerate(
        item_types
    ):
        item_type = (
            raw_item_type or ""
        ).strip().upper()

        raw_item_id = (
            item_ids[index] or ""
        ).strip()

        raw_category_id = (
            category_ids[index] or ""
        ).strip()

        try:
            item_id = int(raw_item_id)
        except (TypeError, ValueError):
            continue

        category_id = None

        if raw_category_id:
            try:
                category_id = int(
                    raw_category_id
                )
            except (TypeError, ValueError):
                flash(
                    "Se recibió una categoría inválida.",
                    "danger",
                )

                return redirect(
                    url_for(
                        "purchases.quotation_request_lines",
                        request_id=request_id,
                        edit_categories=1,
                    )
                )

            submitted_category_ids.add(
                category_id
            )

        if item_type == "ARTICLE":
            if item_id not in valid_article_ids:
                continue

        elif item_type == "PENDING":
            if (
                item_id
                not in valid_pending_article_ids
            ):
                continue

        else:
            continue

        submitted_rows.append(
            {
                "item_type": item_type,
                "item_id": item_id,
                "category_id": category_id,
            }
        )

    # =====================================================
    # 2. VALIDAR TODAS LAS CATEGORÍAS EN UNA CONSULTA
    # =====================================================
    if submitted_category_ids:
        valid_category_ids = {
            row.id
            for row in (
                QuotationCategory.query
                .filter(
                    QuotationCategory.id.in_(
                        submitted_category_ids
                    ),
                    QuotationCategory.is_active.is_(
                        True
                    ),
                )
                .all()
            )
        }

        invalid_category_ids = (
            submitted_category_ids
            - valid_category_ids
        )

        if invalid_category_ids:
            flash(
                "Una de las categorías seleccionadas no existe o está inactiva.",
                "danger",
            )

            return redirect(
                url_for(
                    "purchases.quotation_request_lines",
                    request_id=request_id,
                    edit_categories=1,
                )
            )

    # Evita procesar varias veces el mismo artículo cuando
    # aparece repetido en la misma solicitud.
    normalized_rows = {}

    for item in submitted_rows:
        normalized_rows[
            (
                item["item_type"],
                item["item_id"],
            )
        ] = item

    submitted_rows = list(
        normalized_rows.values()
    )

    article_ids = {
        item["item_id"]
        for item in submitted_rows
        if item["item_type"] == "ARTICLE"
    }

    pending_article_ids = {
        item["item_id"]
        for item in submitted_rows
        if item["item_type"] == "PENDING"
    }

    assignment_filters = []

    if article_ids:
        assignment_filters.append(
            ArticleQuotationCategory.article_id.in_(
                article_ids
            )
        )

    if pending_article_ids:
        assignment_filters.append(
            ArticleQuotationCategory.pending_article_id.in_(
                pending_article_ids
            )
        )

    # =====================================================
    # 3. CARGAR ASIGNACIONES EXISTENTES UNA SOLA VEZ
    # =====================================================
    existing_assignments = []

    if assignment_filters:
        existing_assignments = (
            ArticleQuotationCategory.query
            .filter(
                db.or_(*assignment_filters)
            )
            .all()
        )

    existing_map = {}

    for assignment in existing_assignments:
        if assignment.article_id is not None:
            existing_map[
                (
                    "ARTICLE",
                    assignment.article_id,
                )
            ] = assignment

        elif (
            assignment.pending_article_id
            is not None
        ):
            existing_map[
                (
                    "PENDING",
                    assignment.pending_article_id,
                )
            ] = assignment

    try:
        now = datetime.now(UTC)

        for item in submitted_rows:
            key = (
                item["item_type"],
                item["item_id"],
            )

            category_id = item[
                "category_id"
            ]

            existing = existing_map.get(key)

            # Dejar sin categoría elimina la asignación.
            if category_id is None:
                if existing:
                    db.session.delete(
                        existing
                    )

                continue

            if existing:
                if (
                    existing.quotation_category_id
                    != category_id
                ):
                    existing.quotation_category_id = (
                        category_id
                    )
                    existing.updated_at = now

                continue

            assignment = (
                ArticleQuotationCategory(
                    quotation_category_id=(
                        category_id
                    ),
                    article_id=(
                        item["item_id"]
                        if item["item_type"]
                        == "ARTICLE"
                        else None
                    ),
                    pending_article_id=(
                        item["item_id"]
                        if item["item_type"]
                        == "PENDING"
                        else None
                    ),
                )
            )

            db.session.add(assignment)

        db.session.commit()

    except Exception as exc:
        db.session.rollback()

        print(
            "[SAVE QUOTATION CATEGORIES ERROR] "
            f"{exc}"
        )

        flash(
            "No fue posible guardar las categorías.",
            "danger",
        )

        return redirect(
            url_for(
                "purchases.quotation_request_lines",
                request_id=request_id,
                edit_categories=1,
            )
        )

    flash(
        "Categorías guardadas correctamente.",
        "success",
    )

    return redirect(
        url_for(
            "purchases.quotation_request_lines",
            request_id=request_id,
        )
    )

@purchases_bp.route(
    "/quotations/request/<int:request_id>/export-excel"
)
@login_required
def export_quotation_request_excel(request_id: int):
    purchase_request = PurchaseRequest.query.get_or_404(
        request_id
    )

    # =====================================================
    # 1. CARGAR LÍNEAS DE LA SOLICITUD
    # =====================================================
    request_lines = (
        PurchaseRequestLine.query
        .options(
            joinedload(PurchaseRequestLine.article),
            joinedload(PurchaseRequestLine.pending_article),
            joinedload(PurchaseRequestLine.unit),
        )
        .filter(
            PurchaseRequestLine.purchase_request_id == request_id,
            PurchaseRequestLine.line_status != "CANCELADA",
        )
        .order_by(
            PurchaseRequestLine.id.asc()
        )
        .all()
    )

    line_ids = [
        line.id
        for line in request_lines
    ]

    article_ids = {
        line.article_id
        for line in request_lines
        if line.article_id is not None
    }

    pending_article_ids = {
        line.pending_article_id
        for line in request_lines
        if line.pending_article_id is not None
    }

    # =====================================================
    # 2. CONTAR COTIZACIONES DE TODAS LAS LÍNEAS
    # =====================================================
    quotation_map = {}

    if line_ids:
        quotation_rows = (
            db.session.query(
                QuotationLine.purchase_request_line_id.label(
                    "purchase_request_line_id"
                ),
                func.count(
                    QuotationLine.id
                ).label("total_quotes"),
                func.max(
                    QuotationLine.quote_date
                ).label("last_quote_date"),
            )
            .filter(
                QuotationLine.purchase_request_line_id.in_(
                    line_ids
                )
            )
            .group_by(
                QuotationLine.purchase_request_line_id
            )
            .all()
        )

        quotation_map = {
            row.purchase_request_line_id: row
            for row in quotation_rows
        }

    # =====================================================
    # 3. CARGAR CATEGORÍAS ASIGNADAS EN UNA CONSULTA
    # =====================================================
    assignment_filters = []

    if article_ids:
        assignment_filters.append(
            ArticleQuotationCategory.article_id.in_(
                article_ids
            )
        )

    if pending_article_ids:
        assignment_filters.append(
            ArticleQuotationCategory.pending_article_id.in_(
                pending_article_ids
            )
        )

    article_category_map = {}
    pending_category_map = {}

    if assignment_filters:
        assignment_rows = (
            db.session.query(
                ArticleQuotationCategory.article_id,
                ArticleQuotationCategory.pending_article_id,
                QuotationCategory.id.label(
                    "category_id"
                ),
                QuotationCategory.name.label(
                    "category_name"
                ),
                QuotationCategory.sort_order.label(
                    "category_sort_order"
                ),
            )
            .join(
                QuotationCategory,
                QuotationCategory.id
                == ArticleQuotationCategory.quotation_category_id,
            )
            .filter(
                db.or_(*assignment_filters)
            )
            .all()
        )

        for row in assignment_rows:
            category_name = (
                row.category_name
                or "SIN CATEGORÍA"
            ).replace("¡", "í")

            category_data = {
                "id": row.category_id,
                "name": category_name,
                "sort_order": row.category_sort_order or 999999,
            }

            if row.article_id is not None:
                article_category_map[
                    row.article_id
                ] = category_data

            if row.pending_article_id is not None:
                pending_category_map[
                    row.pending_article_id
                ] = category_data

    # =====================================================
    # 4. AGRUPAR FILAS POR CATEGORÍA
    # =====================================================
    grouped_rows = {}

    for request_line in request_lines:
        category_data = None

        if request_line.article_id is not None:
            category_data = article_category_map.get(
                request_line.article_id
            )

        elif request_line.pending_article_id is not None:
            category_data = pending_category_map.get(
                request_line.pending_article_id
            )

        if category_data:
            category_key = category_data["id"]
            category_name = category_data["name"]
            category_sort_order = category_data["sort_order"]
        else:
            category_key = "SIN_CATEGORIA"
            category_name = "SIN CATEGORÍA"
            category_sort_order = 999999

        quotation_data = quotation_map.get(
            request_line.id
        )

        if category_key not in grouped_rows:
            grouped_rows[category_key] = {
                "name": category_name,
                "sort_order": category_sort_order,
                "rows": [],
            }

        grouped_rows[category_key]["rows"].append(
            {
                "code": request_line.item_code or "-",
                "name": (
                    request_line.item_name
                    or "Sin artículo"
                ),
                "quantity": request_line.quantity_requested,
                "status": (
                    request_line.line_status
                    or "-"
                ).replace("_", " "),
                "total_quotes": int(
                    quotation_data.total_quotes
                    if quotation_data
                    else 0
                ),
                "best_supplier": "-",
                "best_price": None,
                "last_quote_date": (
                    quotation_data.last_quote_date
                    if quotation_data
                    else None
                ),
            }
        )

    groups = sorted(
        grouped_rows.values(),
        key=lambda group: (
            group["sort_order"],
            group["name"],
        ),
    )

    for group in groups:
        group["rows"].sort(
            key=lambda item: (
                item["code"],
                item["name"],
            )
        )

    # =====================================================
    # 5. CREAR ARCHIVO EXCEL
    # =====================================================
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Solicitud de cotización"

    thin_side = Side(
        style="thin",
        color="D1D5DB",
    )

    table_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )

    title_fill = PatternFill(
        "solid",
        fgColor="1D4ED8",
    )

    category_fill = PatternFill(
        "solid",
        fgColor="DBEAFE",
    )

    category_empty_fill = PatternFill(
        "solid",
        fgColor="E2E8F0",
    )

    header_fill = PatternFill(
        "solid",
        fgColor="1E293B",
    )

    sheet.merge_cells("A1:I1")
    sheet["A1"] = "SOLICITUD DE COTIZACIÓN"
    sheet["A1"].font = Font(
        bold=True,
        color="FFFFFF",
        size=16,
    )
    sheet["A1"].fill = title_fill
    sheet["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )
    sheet.row_dimensions[1].height = 26

    sheet["A3"] = "Solicitud:"
    sheet["A3"].font = Font(bold=True)
    sheet["B3"] = purchase_request.number

    sheet["A4"] = "Fecha de exportación:"
    sheet["A4"].font = Font(bold=True)
    sheet["B4"] = datetime.now(CR_TZ).strftime(
        "%d/%m/%Y %H:%M"
    )

    headers = [
        "Código",
        "Artículo",
        "Cantidad",
        "Categoría",
        "Estado",
        "Cotizaciones",
        "Mejor proveedor",
        "Mejor precio",
        "Última fecha",
    ]

    current_row = 6

    for group in groups:
        sheet.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=9,
        )

        category_cell = sheet.cell(
            row=current_row,
            column=1,
            value=(
                f"{group['name']} "
                f"({len(group['rows'])})"
            ),
        )

        category_cell.font = Font(
            bold=True,
            color="0F172A",
            size=12,
        )

        category_cell.fill = (
            category_empty_fill
            if group["name"] == "SIN CATEGORÍA"
            else category_fill
        )

        category_cell.alignment = Alignment(
            vertical="center",
        )

        category_cell.border = table_border
        sheet.row_dimensions[current_row].height = 23

        current_row += 1

        for column_index, header in enumerate(
            headers,
            start=1,
        ):
            cell = sheet.cell(
                row=current_row,
                column=column_index,
                value=header,
            )

            cell.font = Font(
                bold=True,
                color="FFFFFF",
            )

            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
            cell.border = table_border

        current_row += 1

        for item in group["rows"]:
            values = [
                item["code"],
                item["name"],
                float(item["quantity"] or 0),
                group["name"],
                item["status"],
                item["total_quotes"],
                item["best_supplier"],
                (
                    float(item["best_price"])
                    if item["best_price"] is not None
                    else None
                ),
                (
                    item["last_quote_date"].strftime(
                        "%d/%m/%Y"
                    )
                    if item["last_quote_date"]
                    else "-"
                ),
            ]

            for column_index, value in enumerate(
                values,
                start=1,
            ):
                cell = sheet.cell(
                    row=current_row,
                    column=column_index,
                    value=value,
                )

                cell.border = table_border
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                )

                if column_index in [3, 8]:
                    cell.number_format = '#,##0.00'

            current_row += 1

        current_row += 2

    column_widths = {
        "A": 17,
        "B": 52,
        "C": 14,
        "D": 22,
        "E": 23,
        "F": 15,
        "G": 28,
        "H": 17,
        "I": 17,
    }

    for column_letter, width in column_widths.items():
        sheet.column_dimensions[
            column_letter
        ].width = width

    sheet.freeze_panes = "A6"
    sheet.sheet_view.showGridLines = False

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    safe_request_number = (
        purchase_request.number
        or str(request_id)
    ).replace("/", "-")

    return send_file(
        output,
        as_attachment=True,
        download_name=(
            f"solicitud_cotizacion_"
            f"{safe_request_number}.xlsx"
        ),
        mimetype=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        ),
    )

@purchases_bp.route(
    "/quotations/create",
    methods=["GET", "POST"],
)
@login_required
def create_quotation():
    purchase_requests = (
        PurchaseRequest.query
        .options(
            joinedload(PurchaseRequest.site),
            joinedload(PurchaseRequest.warehouse),
        )
        .filter(
            PurchaseRequest.status.in_(
                [
                    "EN_REVISION_PROVEEDURIA",
                    "PARCIALMENTE_COTIZADA",
                    "COTIZADA",
                ]
            )
        )
        .order_by(
            PurchaseRequest.created_at.desc(),
            PurchaseRequest.id.desc(),
        )
        .limit(100)
        .all()
    )

    return render_template(
        "purchases/quotations/create.html",
        purchase_requests=purchase_requests,
    )

@purchases_bp.route(
    "/quotations/request/<int:request_id>/partial-lines"
)
@login_required
def quotation_request_lines_partial(request_id: int):

    rows = (
        db.session.query(
            PurchaseRequestLine
        )
        .options(
            joinedload(PurchaseRequestLine.unit)
        )
        .filter(
            PurchaseRequestLine.purchase_request_id == request_id,
            PurchaseRequestLine.line_status.notin_([
                "CANCELADA",
                "CONVERTIDA_A_OC",
                "RECIBIDA",
            ])
        )
        .order_by(PurchaseRequestLine.id.asc())
        .all()
    )

    return render_template(
        "purchases/quotations/_request_lines.html",
        lines=rows,
    )


@purchases_bp.route("/quotations/<int:batch_id>")
@login_required
def quotation_detail(batch_id: int):
    quotation_batch = get_quotation_batch_or_404(batch_id)
    return render_template(
        "purchases/quotations/detail.html",
        quotation_batch=quotation_batch,
    )

@purchases_bp.route(
    "/quotations/free/article",
    methods=["GET", "POST"]
)
@login_required
def quotation_free_article():

    suppliers = (
        Supplier.query
        .filter(Supplier.is_active.is_(True))
        .order_by(
            Supplier.commercial_name.asc()
        )
        .limit(300)
        .all()
    )

    return render_template(
        "purchases/quotations/free_article.html",
        suppliers=suppliers,
    )

@purchases_bp.route(
    "/quotations/articles/search"
)
@login_required
def quotation_articles_search():

    q = (
        request.args.get("q") or ""
    ).strip()

    if len(q) < 2:
        return {"items": []}

    rows = (
        Article.query
        .filter(
            Article.is_active.is_(True),
            db.or_(
                Article.code.ilike(f"%{q}%"),
                Article.name.ilike(f"%{q}%"),
            )
        )
        .order_by(
            Article.code.asc()
        )
        .limit(20)
        .all()
    )

    return {
        "items": [
            {
                "id": row.id,
                "code": row.code,
                "name": row.name,
            }
            for row in rows
        ]
    }

@purchases_bp.route("/quotations/free", methods=["GET", "POST"])
@login_required
def quotation_free():
    if request.method == "POST":
        article_id = _to_int(request.form.get("article_id"))
        new_article_name = (request.form.get("new_article_name") or "").strip()

        supplier_id = _to_int(request.form.get("supplier_id"))
        new_supplier_name = (request.form.get("new_supplier_name") or "").strip()

        quote_date = request.form.get("quote_date")
        unit_price = request.form.get("unit_price")
        currency_code = request.form.get("currency_code") or "CRC"
        discount_pct = request.form.get("discount_pct") or "0"
        tax_pct = request.form.get("tax_pct") or "0"
        tax_included = request.form.get("tax_included") == "1"
        lead_time_days = request.form.get("lead_time_days")
        brand_model = request.form.get("brand_model")
        notes = request.form.get("notes")
        payment_type = request.form.get("payment_type") or None
        payment_term_months = request.form.get("payment_term_months")
        origin_type = request.form.get("origin_type") or None

        pending_article_id = None

        try:
            if not article_id and not new_article_name:
                raise QuotationServiceError(
                    "Debe seleccionar un artículo existente o escribir un artículo nuevo."
                )

            if article_id and new_article_name:
                raise QuotationServiceError(
                    "Debe usar artículo existente o artículo nuevo, no ambos."
                )

            if new_article_name:
                pending = create_pending_article(
                    provisional_name=new_article_name,
                    description=notes,
                    category_id=None,
                    unit_id=None,
                    requested_by_user_id=current_user.id,
                )
                pending_article_id = pending.id

            if new_supplier_name:
                supplier = create_minimal_supplier_for_quotation(
                    commercial_name=new_supplier_name,
                )
                supplier_id = supplier.id

            if not supplier_id:
                raise QuotationServiceError("Debe seleccionar o crear un proveedor.")

            line = QuotationLinePayload(
                purchase_request_line_id=None,
                supplier_id=supplier_id,
                quote_date=quote_date,
                unit_price=_to_decimal(unit_price),
                currency_code=currency_code,
                article_id=article_id,
                pending_article_id=pending_article_id,
                discount_pct=_to_decimal(discount_pct),
                tax_pct=_to_decimal(tax_pct),
                tax_included=tax_included,
                lead_time_days=_to_int(lead_time_days),
                brand_model=brand_model,
                notes=notes,
                payment_type=payment_type,
                payment_term_months=_to_int(payment_term_months),
                origin_type=origin_type,
                status="COTIZADA",
            )

            quotation_batch = create_quotation_batch(
                purchase_request_id=None,
                created_by_user_id=current_user.id,
                quote_date=quote_date,
                notes=notes,
                lines=[line],
            )

            flash("Cotización libre creada correctamente.", "success")
            return redirect(
                url_for(
                    "purchases.quotation_detail",
                    batch_id=quotation_batch.id,
                )
            )

        except QuotationServiceError as exc:
            flash(str(exc), "danger")

        except Exception as exc:
            print(f"[FREE QUOTATION ERROR] {exc}")
            db.session.rollback()
            flash("Error interno al crear la cotización libre.", "danger")

    return render_template(
        "purchases/quotations/free.html",
    )

@purchases_bp.route("/suppliers/search")
@login_required
def search_suppliers_for_quotation():
    q = (request.args.get("q") or "").strip()

    if len(q) < 2:
        return {"items": []}

    like_value = f"%{q}%"

    suppliers = (
        Supplier.query
        .filter(
            Supplier.is_active.is_(True),
            db.or_(
                Supplier.commercial_name.ilike(like_value),
                Supplier.legal_name.ilike(like_value),
                Supplier.tax_id.ilike(like_value),
            )
        )
        .order_by(
            Supplier.commercial_name.asc(),
            Supplier.legal_name.asc(),
        )
        .limit(20)
        .all()
    )

    return {
        "items": [
            {
                "id": supplier.id,
                "commercial_name": supplier.commercial_name or "",
                "legal_name": supplier.legal_name or "",
                "tax_id": supplier.tax_id or "",
            }
            for supplier in suppliers
        ]
    }
# =========================
# ORDENES DE COMPRA
# =========================
@purchases_bp.route("/orders")
@login_required
def list_orders():
    approval_status = request.args.get("approval_status", type=str)
    supplier_id = _to_int(request.args.get("supplier_id"))
    search = request.args.get("search", type=str)

    return render_template(
        "purchases/orders/index.html",
        purchase_orders=[],
        suppliers=[],
        pagination=None,
        selected_approval_status=approval_status,
        selected_supplier_id=supplier_id,
        search=search,
    )

@purchases_bp.route("/orders/partial/list")
@login_required
def list_orders_partial():
    approval_status = request.args.get("approval_status", type=str)
    supplier_id = _to_int(request.args.get("supplier_id"))
    search = request.args.get("search", type=str)
    page = request.args.get("page", 1, type=int)

    try:
        query = (
            PurchaseOrder.query
            .options(
                joinedload(PurchaseOrder.supplier),
                joinedload(PurchaseOrder.purchase_request),
                joinedload(PurchaseOrder.warehouse),
            )
        )

        if approval_status:
            query = query.filter(
                PurchaseOrder.approval_status == approval_status
            )

        if supplier_id:
            query = query.filter(
                PurchaseOrder.supplier_id == supplier_id
            )

        if search:
            like_value = f"%{search.strip()}%"
            query = query.filter(
                PurchaseOrder.number.ilike(like_value)
            )

        pagination = (
            query
            .order_by(
                PurchaseOrder.created_at.desc(),
                PurchaseOrder.id.desc(),
            )
            .paginate(
                page=page,
                per_page=20,
                error_out=False,
            )
        )

        return render_template(
            "purchases/orders/_list.html",
            purchase_orders=pagination.items,
            pagination=pagination,
            selected_approval_status=approval_status,
            selected_supplier_id=supplier_id,
            search=search,
        )

    except Exception as exc:
        print(f"[PURCHASE ORDERS PARTIAL ERROR] {exc}")
        db.session.rollback()

        return render_template(
            "purchases/orders/_list.html",
            purchase_orders=[],
            pagination=None,
            selected_approval_status=approval_status,
            selected_supplier_id=supplier_id,
            search=search,
        ), 500


@purchases_bp.route(
    "/orders/create",
    methods=["GET", "POST"],
)
@login_required
def create_order():
    """
    Muestra las selecciones pendientes de OC y genera una orden
    exclusivamente desde PurchaseOrderCandidate.

    Ya no recibe QuotationLine directamente ni construye manualmente
    PurchaseOrderLinePayload.
    """

    selected_supplier_id = _to_int(
        request.args.get("supplier_id")
    )

    if request.method == "POST":
        raw_candidate_ids = request.form.getlist(
            "candidate_id[]"
        )

        candidate_ids: list[int] = []

        for raw_candidate_id in raw_candidate_ids:
            candidate_id = _to_int(
                raw_candidate_id
            )

            if candidate_id:
                candidate_ids.append(
                    candidate_id
                )

        supplier_id = _to_int(
            request.form.get("supplier_id")
        )

        payment_terms = (
            request.form.get("payment_terms")
            or ""
        ).strip() or None

        notes = (
            request.form.get("notes")
            or ""
        ).strip() or None

        if not candidate_ids:
            flash(
                "Debe seleccionar al menos una línea pendiente para crear la orden de compra.",
                "danger",
            )

            return redirect(
                url_for(
                    "purchases.create_order",
                    supplier_id=supplier_id,
                )
            )

        try:
            purchase_order = create_purchase_order(
                candidate_ids=candidate_ids,
                generated_by_user_id=current_user.id,
                supplier_id=supplier_id,
                payment_terms=payment_terms,
                notes=notes,
            )

        except PurchaseOrderServiceError as exc:
            flash(
                str(exc),
                "danger",
            )

            return redirect(
                url_for(
                    "purchases.create_order",
                    supplier_id=supplier_id,
                )
            )

        flash(
            "Orden de compra creada correctamente y enviada a aprobación.",
            "success",
        )

        return redirect(
            url_for(
                "purchases.order_print",
                order_id=purchase_order.id,
            )
        )

    return render_template(
        "purchases/orders/create.html",
        candidate_groups=[],
        selected_supplier_id=(
            selected_supplier_id
        ),
    )

@purchases_bp.route(
    "/orders/create/partial/quotation-groups"
)
@login_required
def create_order_quotation_groups_partial():
    """
    Devuelve candidatos pendientes de orden de compra agrupados
    por proveedor.

    La URL se conserva para no romper las llamadas existentes del
    frontend, aunque la fuente actual es PurchaseOrderCandidate.
    """

    supplier_id = _to_int(
        request.args.get("supplier_id")
    )

    search = (
        request.args.get("search")
        or ""
    ).strip().lower()

    try:
        candidate_groups = (
            list_pending_purchase_order_candidates_by_supplier(
                supplier_id=supplier_id,
            )
        )

        if search:
            filtered_groups = []

            for group in candidate_groups:
                supplier_name = (
                    group.get("supplier_name")
                    or ""
                ).lower()

                matching_lines = []

                for line in group.get(
                    "lines",
                    [],
                ):
                    searchable_values = [
                        supplier_name,
                        str(
                            line.get(
                                "purchase_request_number"
                            )
                            or ""
                        ).lower(),
                        str(
                            line.get("item_code")
                            or ""
                        ).lower(),
                        str(
                            line.get("item_name")
                            or ""
                        ).lower(),
                        str(
                            line.get(
                                "quotation_category_label"
                            )
                            or ""
                        ).lower(),
                        str(
                            line.get("currency_code")
                            or ""
                        ).lower(),
                        str(
                            line.get("notes")
                            or ""
                        ).lower(),
                    ]

                    if any(
                        search in searchable_value
                        for searchable_value
                        in searchable_values
                    ):
                        matching_lines.append(
                            line
                        )

                if not matching_lines:
                    continue

                filtered_group = dict(
                    group
                )

                filtered_group["lines"] = (
                    matching_lines
                )

                filtered_group["total_lines"] = (
                    len(matching_lines)
                )

                currencies = sorted(
                    {
                        line.get(
                            "currency_code"
                        )
                        or "CRC"
                        for line in matching_lines
                    }
                )

                totals_by_currency_map: dict[
                    str,
                    dict,
                ] = {}

                for line in matching_lines:
                    currency_code = (
                        line.get(
                            "currency_code"
                        )
                        or "CRC"
                    )

                    currency_totals = (
                        totals_by_currency_map.setdefault(
                            currency_code,
                            {
                                "currency_code": (
                                    currency_code
                                ),
                                "subtotal": Decimal(
                                    "0"
                                ),
                                "discount_amount": Decimal(
                                    "0"
                                ),
                                "tax_amount": Decimal(
                                    "0"
                                ),
                                "total": Decimal(
                                    "0"
                                ),
                                "total_lines": 0,
                            },
                        )
                    )

                    currency_totals[
                        "subtotal"
                    ] += Decimal(
                        str(
                            line.get(
                                "line_subtotal"
                            )
                            or 0
                        )
                    )

                    currency_totals[
                        "discount_amount"
                    ] += Decimal(
                        str(
                            line.get(
                                "discount_amount"
                            )
                            or 0
                        )
                    )

                    currency_totals[
                        "tax_amount"
                    ] += Decimal(
                        str(
                            line.get(
                                "tax_amount"
                            )
                            or 0
                        )
                    )

                    currency_totals[
                        "total"
                    ] += Decimal(
                        str(
                            line.get(
                                "line_total"
                            )
                            or 0
                        )
                    )

                    currency_totals[
                        "total_lines"
                    ] += 1

                filtered_group[
                    "currencies"
                ] = currencies

                filtered_group[
                    "has_mixed_currencies"
                ] = len(currencies) > 1

                filtered_group[
                    "currency_code"
                ] = (
                    currencies[0]
                    if len(currencies) == 1
                    else None
                )

                filtered_group[
                    "totals_by_currency"
                ] = [
                    totals_by_currency_map[
                        currency_code
                    ]
                    for currency_code in currencies
                ]

                if len(currencies) == 1:
                    only_total = (
                        filtered_group[
                            "totals_by_currency"
                        ][0]
                    )

                    filtered_group[
                        "subtotal"
                    ] = only_total["subtotal"]

                    filtered_group[
                        "discount_amount"
                    ] = only_total[
                        "discount_amount"
                    ]

                    filtered_group[
                        "tax_amount"
                    ] = only_total[
                        "tax_amount"
                    ]

                    filtered_group[
                        "total"
                    ] = only_total["total"]

                else:
                    filtered_group[
                        "subtotal"
                    ] = None

                    filtered_group[
                        "discount_amount"
                    ] = None

                    filtered_group[
                        "tax_amount"
                    ] = None

                    filtered_group[
                        "total"
                    ] = None

                filtered_groups.append(
                    filtered_group
                )

            candidate_groups = filtered_groups

        return {
            "items": candidate_groups,
        }

    except QuotationServiceError as exc:
        return {
            "items": [],
            "error": str(exc),
        }, 400

    except Exception as exc:
        db.session.rollback()

        print(
            "[PURCHASE ORDER CANDIDATES PARTIAL ERROR] "
            f"{exc}"
        )

        return {
            "items": [],
            "error": (
                "No fue posible cargar las líneas pendientes de orden de compra."
            ),
        }, 500

@purchases_bp.route(
    "/orders/create/manual",
    methods=["GET", "POST"],
)
@login_required
def create_order_manual():
    """
    Crea una orden de compra manual sin solicitud de compra,
    cotización ni candidato de pre-OC.

    La orden se crea directamente en:

    - atm.purchase_orders
    - atm.purchase_order_lines

    El flujo normal desde cotización permanece intacto.
    """

    suppliers = (
        Supplier.query
        .filter(
            Supplier.is_active.is_(True)
        )
        .order_by(
            Supplier.commercial_name.asc(),
            Supplier.legal_name.asc(),
            Supplier.id.asc(),
        )
        .limit(500)
        .all()
    )

    sites = (
        Site.query
        .filter(
            Site.is_active.is_(True)
        )
        .order_by(
            Site.name.asc(),
            Site.id.asc(),
        )
        .all()
    )

    warehouses = (
        Warehouse.query
        .filter(
            Warehouse.is_active.is_(True)
        )
        .order_by(
            Warehouse.name.asc(),
            Warehouse.id.asc(),
        )
        .all()
    )

    units = (
        Unit.query
        .order_by(
            Unit.id.asc()
        )
        .all()
    )

    active_site_id = session.get(
        "active_site_id"
    )

    submitted_header = {
        "supplier_id": "",
        "site_id": (
            str(active_site_id)
            if active_site_id
            else ""
        ),
        "warehouse_id": "",
        "currency_code": "CRC",
        "payment_terms": "",
        "notes": "",
    }

    submitted_lines = [
        {
            "article_id": "",
            "article_search": "",
            "pending_article_id": "",
            "quantity": "",
            "unit_id": "",
            "unit_cost": "",
            "discount_pct": "0",
            "tax_pct": "13",
            "notes": "",
        }
    ]

    if request.method == "GET":
        return render_template(
            "purchases/orders/create_manual.html",
            suppliers=suppliers,
            sites=sites,
            warehouses=warehouses,
            units=units,
            submitted_header=submitted_header,
            submitted_lines=submitted_lines,
        )

    # =====================================================
    # 1. RECIBIR ENCABEZADO
    # =====================================================

    supplier_id_raw = (
        request.form.get("supplier_id")
        or ""
    ).strip()

    site_id_raw = (
        request.form.get("site_id")
        or ""
    ).strip()

    warehouse_id_raw = (
        request.form.get("warehouse_id")
        or ""
    ).strip()

    currency_code = (
        request.form.get("currency_code")
        or "CRC"
    ).strip().upper()

    payment_terms = (
        request.form.get("payment_terms")
        or ""
    ).strip()

    notes = (
        request.form.get("notes")
        or ""
    ).strip()

    submitted_header = {
        "supplier_id": supplier_id_raw,
        "site_id": site_id_raw,
        "warehouse_id": warehouse_id_raw,
        "currency_code": currency_code,
        "payment_terms": payment_terms,
        "notes": notes,
    }

    # =====================================================
    # 2. RECIBIR LÍNEAS
    # =====================================================

    article_ids = request.form.getlist(
        "line_article_id[]"
    )

    article_searches = request.form.getlist(
        "line_article_search[]"
    )

    pending_article_ids = request.form.getlist(
        "line_pending_article_id[]"
    )

    quantities = request.form.getlist(
        "line_quantity[]"
    )

    unit_ids = request.form.getlist(
        "line_unit_id[]"
    )

    unit_costs = request.form.getlist(
        "line_unit_cost[]"
    )

    discount_pcts = request.form.getlist(
        "line_discount_pct[]"
    )

    tax_pcts = request.form.getlist(
        "line_tax_pct[]"
    )

    line_notes_list = request.form.getlist(
        "line_notes[]"
    )

    max_len = max(
        [
            len(article_ids),
            len(article_searches),
            len(pending_article_ids),
            len(quantities),
            len(unit_ids),
            len(unit_costs),
            len(discount_pcts),
            len(tax_pcts),
            len(line_notes_list),
        ],
        default=0,
    )

    submitted_lines = []

    for index in range(max_len):
        article_id_raw = (
            article_ids[index].strip()
            if index < len(article_ids)
            else ""
        )

        article_search_raw = (
            article_searches[index].strip()
            if index < len(article_searches)
            else ""
        )

        pending_article_id_raw = (
            pending_article_ids[index].strip()
            if index < len(pending_article_ids)
            else ""
        )

        quantity_raw = (
            quantities[index].strip()
            if index < len(quantities)
            else ""
        )

        unit_id_raw = (
            unit_ids[index].strip()
            if index < len(unit_ids)
            else ""
        )

        unit_cost_raw = (
            unit_costs[index].strip()
            if index < len(unit_costs)
            else ""
        )

        discount_pct_raw = (
            discount_pcts[index].strip()
            if index < len(discount_pcts)
            else "0"
        )

        tax_pct_raw = (
            tax_pcts[index].strip()
            if index < len(tax_pcts)
            else "13"
        )

        line_notes = (
            line_notes_list[index].strip()
            if index < len(line_notes_list)
            else ""
        )

        row = {
            "article_id": article_id_raw,
            "article_search": article_search_raw,
            "pending_article_id": pending_article_id_raw,
            "quantity": quantity_raw,
            "unit_id": unit_id_raw,
            "unit_cost": unit_cost_raw,
            "discount_pct": (
                discount_pct_raw or "0"
            ),
            "tax_pct": (
                tax_pct_raw or "0"
            ),
            "notes": line_notes,
        }

        # Ignorar únicamente filas totalmente vacías.
        if not any(
            [
                article_id_raw,
                article_search_raw,
                pending_article_id_raw,
                quantity_raw,
                unit_id_raw,
                unit_cost_raw,
                line_notes,
            ]
        ):
            continue

        submitted_lines.append(
            row
        )

    if not submitted_lines:
        submitted_lines = [
            {
                "article_id": "",
                "article_search": "",
                "pending_article_id": "",
                "quantity": "",
                "unit_id": "",
                "unit_cost": "",
                "discount_pct": "0",
                "tax_pct": "13",
                "notes": "",
            }
        ]

        flash(
            "Debe agregar al menos un artículo a la orden.",
            "danger",
        )

        return render_template(
            "purchases/orders/create_manual.html",
            suppliers=suppliers,
            sites=sites,
            warehouses=warehouses,
            units=units,
            submitted_header=submitted_header,
            submitted_lines=submitted_lines,
        )

    # =====================================================
    # 3. NORMALIZAR ENCABEZADO
    # =====================================================

    try:
        supplier_id = (
            int(supplier_id_raw)
            if supplier_id_raw
            else None
        )

        site_id = (
            int(site_id_raw)
            if site_id_raw
            else None
        )

        warehouse_id = (
            int(warehouse_id_raw)
            if warehouse_id_raw
            else None
        )

    except ValueError:
        flash(
            "El proveedor, predio o bodega seleccionado no es válido.",
            "danger",
        )

        return render_template(
            "purchases/orders/create_manual.html",
            suppliers=suppliers,
            sites=sites,
            warehouses=warehouses,
            units=units,
            submitted_header=submitted_header,
            submitted_lines=submitted_lines,
        )

    if not supplier_id:
        flash(
            "Debe seleccionar un proveedor.",
            "danger",
        )

        return render_template(
            "purchases/orders/create_manual.html",
            suppliers=suppliers,
            sites=sites,
            warehouses=warehouses,
            units=units,
            submitted_header=submitted_header,
            submitted_lines=submitted_lines,
        )

    if not warehouse_id:
        flash(
            "Debe seleccionar la bodega que recibirá la compra.",
            "danger",
        )

        return render_template(
            "purchases/orders/create_manual.html",
            suppliers=suppliers,
            sites=sites,
            warehouses=warehouses,
            units=units,
            submitted_header=submitted_header,
            submitted_lines=submitted_lines,
        )

    # =====================================================
    # 4. CONSTRUIR PAYLOAD PARA EL SERVICIO
    # =====================================================

    lines = []

    for index, row in enumerate(
        submitted_lines,
        start=1,
    ):
        try:
            article_id = (
                int(row["article_id"])
                if row["article_id"]
                else None
            )

            pending_article_id = (
                int(row["pending_article_id"])
                if row["pending_article_id"]
                else None
            )

            unit_id = (
                int(row["unit_id"])
                if row["unit_id"]
                else None
            )

        except ValueError:
            flash(
                f"Los datos del artículo en la línea {index} no son válidos.",
                "danger",
            )

            return render_template(
                "purchases/orders/create_manual.html",
                suppliers=suppliers,
                sites=sites,
                warehouses=warehouses,
                units=units,
                submitted_header=submitted_header,
                submitted_lines=submitted_lines,
            )

        if not article_id and not pending_article_id:
            flash(
                f"Debe seleccionar un artículo en la línea {index}.",
                "danger",
            )

            return render_template(
                "purchases/orders/create_manual.html",
                suppliers=suppliers,
                sites=sites,
                warehouses=warehouses,
                units=units,
                submitted_header=submitted_header,
                submitted_lines=submitted_lines,
            )

        if article_id and pending_article_id:
            flash(
                f"La línea {index} no puede contener un artículo "
                "existente y uno pendiente al mismo tiempo.",
                "danger",
            )

            return render_template(
                "purchases/orders/create_manual.html",
                suppliers=suppliers,
                sites=sites,
                warehouses=warehouses,
                units=units,
                submitted_header=submitted_header,
                submitted_lines=submitted_lines,
            )

        lines.append(
            {
                "article_id": article_id,
                "pending_article_id": (
                    pending_article_id
                ),
                "quantity": row["quantity"],
                "unit_id": unit_id,
                "unit_cost": row["unit_cost"],
                "discount_pct": (
                    row["discount_pct"]
                    or "0"
                ),
                "tax_pct": (
                    row["tax_pct"]
                    or "0"
                ),
                "notes": (
                    row["notes"]
                    or None
                ),
            }
        )

    # =====================================================
    # 5. CREAR OC MANUAL
    # =====================================================

    try:
        purchase_order = (
            create_manual_purchase_order(
                supplier_id=supplier_id,
                generated_by_user_id=(
                    current_user.id
                ),
                warehouse_id=warehouse_id,
                site_id=site_id,
                currency_code=currency_code,
                payment_terms=(
                    payment_terms
                    or None
                ),
                notes=notes or None,
                lines=lines,
            )
        )

    except PurchaseOrderServiceError as exc:
        db.session.rollback()

        flash(
            str(exc),
            "danger",
        )

        flash(
            "La orden no fue creada. "
            "Los datos ingresados se conservaron.",
            "warning",
        )

        return render_template(
            "purchases/orders/create_manual.html",
            suppliers=suppliers,
            sites=sites,
            warehouses=warehouses,
            units=units,
            submitted_header=submitted_header,
            submitted_lines=submitted_lines,
        )

    except Exception as exc:
        db.session.rollback()

        print(
            "[CREATE MANUAL PURCHASE ORDER ROUTE ERROR] "
            f"{type(exc).__name__}: {exc}"
        )

        flash(
            "Ocurrió un error inesperado al crear "
            "la orden de compra manual.",
            "danger",
        )

        return render_template(
            "purchases/orders/create_manual.html",
            suppliers=suppliers,
            sites=sites,
            warehouses=warehouses,
            units=units,
            submitted_header=submitted_header,
            submitted_lines=submitted_lines,
        )

    flash(
        "Orden de compra manual creada correctamente "
        "y enviada a aprobación.",
        "success",
    )

    return redirect(
        url_for(
            "purchases.order_print",
            order_id=purchase_order.id,
        )
    )

@purchases_bp.route("/orders/<int:order_id>")
@login_required
def order_detail(order_id: int):
    purchase_order = get_purchase_order_or_404(order_id)
    return render_template(
        "purchases/orders/detail.html",
        purchase_order=purchase_order,
        cr_datetime=_cr_datetime,
    )


@purchases_bp.route("/orders/<int:order_id>/print")
@login_required
def order_print(order_id: int):
    purchase_order = get_purchase_order_or_404(order_id)
    return render_template(
        "purchases/orders/print.html",
        purchase_order=purchase_order,
        cr_datetime=_cr_datetime,
    )


@purchases_bp.route("/orders/<int:order_id>/approve", methods=["POST"])
@login_required
def approve_order(order_id: int):
    try:
        register_purchase_order_approval(
            purchase_order_id=order_id,
            approved_by_user_id=current_user.id,
            status="APROBADA",
            reason=request.form.get("reason"),
        )
    except PurchaseOrderServiceError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("purchases.order_detail", order_id=order_id))

    flash("Orden de compra aprobada.", "success")
    return redirect(url_for("purchases.order_detail", order_id=order_id))


@purchases_bp.route("/orders/<int:order_id>/reject", methods=["POST"])
@login_required
def reject_order(order_id: int):
    try:
        register_purchase_order_approval(
            purchase_order_id=order_id,
            approved_by_user_id=current_user.id,
            status="RECHAZADA",
            reason=request.form.get("reason"),
        )
    except PurchaseOrderServiceError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("purchases.order_detail", order_id=order_id))

    flash("Orden de compra rechazada.", "warning")
    return redirect(url_for("purchases.order_detail", order_id=order_id))


# =========================
# ENTRADAS A INVENTARIO
# =========================
@purchases_bp.route("/inventory-entries")
@login_required
def list_entries():
    search = request.args.get("search", type=str)

    return render_template(
        "purchases/inventory_entries/index.html",
        inventory_entries=[],
        pagination=None,
        search=search,
    )

@purchases_bp.route("/inventory-entries/partial/list")
@login_required
def list_entries_partial():
    search = request.args.get("search", type=str)
    page = request.args.get("page", 1, type=int)

    try:
        query = (
            InventoryEntry.query
            .options(
                joinedload(InventoryEntry.purchase_order),
                joinedload(InventoryEntry.warehouse),
                joinedload(InventoryEntry.entered_by_user),
            )
        )

        if search:
            like_value = f"%{search.strip()}%"
            query = query.filter(
                db.or_(
                    InventoryEntry.number.ilike(like_value),
                    InventoryEntry.invoice_number.ilike(like_value),
                )
            )

        pagination = (
            query
            .order_by(
                InventoryEntry.created_at.desc(),
                InventoryEntry.id.desc(),
            )
            .paginate(
                page=page,
                per_page=15,
                error_out=False,
            )
        )

        return render_template(
            "purchases/inventory_entries/_list.html",
            inventory_entries=pagination.items,
            pagination=pagination,
            search=search,
        )

    except Exception as exc:
        print(f"[INVENTORY ENTRIES PARTIAL ERROR] {exc}")
        db.session.rollback()

        return render_template(
            "purchases/inventory_entries/_list.html",
            inventory_entries=[],
            pagination=None,
            search=search,
        ), 500

@purchases_bp.route("/entries/create", methods=["GET", "POST"])
@login_required
def create_entry():

    purchase_orders = _get_valid_purchase_orders_for_receiving()

    warehouses = (
        Warehouse.query
        .filter_by(is_active=True)
        .order_by(Warehouse.name.asc())
        .all()
    )

    units = (
        Unit.query
        .order_by(Unit.id.asc())
        .all()
    )

    if request.method == "POST":
        purchase_order_id = _to_int(request.form.get("purchase_order_id"))
        supplier_id = _to_int(request.form.get("supplier_id"))
        warehouse_id = _to_int(request.form.get("warehouse_id"))
        invoice_number = request.form.get("invoice_number")
        invoice_date = request.form.get("invoice_date")
        notes = request.form.get("notes")

        po_line_ids = request.form.getlist("line_purchase_order_line_id[]")
        location_ids = request.form.getlist("line_warehouse_location_id[]")
        article_ids = request.form.getlist("line_article_id[]")
        pending_ids = request.form.getlist("line_pending_article_id[]")
        quantities = request.form.getlist("line_quantity_received[]")
        unit_ids = request.form.getlist("line_unit_id[]")
        cost_wo_tax = request.form.getlist("line_unit_cost_without_tax[]")
        cost_w_tax = request.form.getlist("line_unit_cost_with_tax[]")
        discounts = request.form.getlist("line_discount_pct[]")
        taxes = request.form.getlist("line_tax_pct[]")
        line_notes_list = request.form.getlist("line_notes[]")

        max_len = max(
            [
                len(article_ids),
                len(pending_ids),
                len(quantities),
            ],
            default=0,
        )

        lines: list[InventoryEntryLinePayload] = []

        for i in range(max_len):
            quantity_raw = quantities[i] if i < len(quantities) else None
            article_id = _to_int(article_ids[i] if i < len(article_ids) else None)
            pending_id = _to_int(pending_ids[i] if i < len(pending_ids) else None)
            po_line_id = _to_int(po_line_ids[i] if i < len(po_line_ids) else None)

            if not any([quantity_raw, article_id, pending_id, po_line_id]):
                continue

            try:
                quantity = _to_decimal(quantity_raw)
                cost1 = _to_decimal(cost_wo_tax[i] if i < len(cost_wo_tax) else None)
                cost2 = _to_decimal(cost_w_tax[i] if i < len(cost_w_tax) else None)
                discount = _to_decimal(discounts[i] if i < len(discounts) else None)
                tax = _to_decimal(taxes[i] if i < len(taxes) else None)

            except ValueError:
                flash(f"Error en línea {i + 1}", "danger")

                return render_template(
                    "purchases/inventory_entries/create.html",
                    purchase_orders=purchase_orders,
                    warehouses=warehouses,
                    units=units,
                )

            lines.append(
                InventoryEntryLinePayload(
                    purchase_order_line_id=po_line_id,
                    article_id=article_id,
                    pending_article_id=pending_id,
                    warehouse_location_id=_to_int(
                        location_ids[i] if i < len(location_ids) else None
                    ),
                    quantity_received=quantity,
                    unit_id=_to_int(unit_ids[i] if i < len(unit_ids) else None),
                    unit_cost_without_tax=cost1,
                    unit_cost_with_tax=cost2,
                    discount_pct=discount,
                    tax_pct=tax,
                    line_notes=line_notes_list[i] if i < len(line_notes_list) else None,
                )
            )

        try:
            entry = create_inventory_entry(
                purchase_order_id=purchase_order_id,
                supplier_id=supplier_id,
                warehouse_id=warehouse_id,
                entered_by_user_id=current_user.id,
                invoice_number=invoice_number,
                invoice_date=invoice_date,
                notes=notes,
                lines=lines,
            )

        except InventoryEntryServiceError as exc:
            flash(str(exc), "danger")

            return render_template(
                "purchases/inventory_entries/create.html",
                purchase_orders=purchase_orders,
                warehouses=warehouses,
                units=units,
            )

        flash("Entrada registrada correctamente.", "success")
        return redirect(url_for("purchases.entry_detail", entry_id=entry.id))

    return render_template(
        "purchases/inventory_entries/create.html",
        purchase_orders=purchase_orders,
        warehouses=warehouses,
        units=units,
    )

# =========================================================
# AJAX - LÍNEAS DE OC PARA ENTRADAS
# =========================================================
@purchases_bp.route("/orders/<int:order_id>/lines-for-entry")
@login_required
def get_order_lines_for_entry(order_id):

    try:

        purchase_order = (
            PurchaseOrder.query
            .options(
                joinedload(PurchaseOrder.lines)
                .joinedload(PurchaseOrderLine.unit),

                joinedload(PurchaseOrder.lines)
                .joinedload(PurchaseOrderLine.article),

                joinedload(PurchaseOrder.lines)
                .joinedload(PurchaseOrderLine.pending_article),
            )
            .get_or_404(order_id)
        )

        items = []

        for line in purchase_order.lines:

            quantity_ordered = float(line.quantity_ordered or 0)
            quantity_received = float(line.quantity_received or 0)

            pending_quantity = max(
                quantity_ordered - quantity_received,
                0,
            )

            items.append({
                "id": line.id,

                "item_name": line.item_name or "Sin artículo",

                "item_code": line.item_code or "",

                "article_id": line.article_id,

                "pending_article_id": line.pending_article_id,

                "quantity_ordered": quantity_ordered,

                "quantity_received": quantity_received,

                "pending_quantity": pending_quantity,

                "unit_id": line.unit_id,

                "unit_name": (
                    line.unit.name
                    if line.unit else ""
                ),

                "unit_cost": float(line.unit_cost or 0),

                "discount_pct": float(line.discount_pct or 0),

                "tax_pct": float(line.tax_pct or 0),
            })

        return {
            "items": items
        }

    except Exception as exc:

        print(f"[ORDER LINES FOR ENTRY ERROR] {exc}")

        return {
            "items": []
        }, 500


# =========================================================
# AJAX - UBICACIONES DE BODEGA
# =========================================================
@purchases_bp.route("/warehouses/<int:warehouse_id>/locations")
@login_required
def get_warehouse_locations(warehouse_id):

    try:

        locations = (
            WarehouseLocation.query
            .filter(
                WarehouseLocation.warehouse_id == warehouse_id,
                WarehouseLocation.is_active.is_(True),
            )
            .order_by(
                WarehouseLocation.code.asc(),
                WarehouseLocation.id.asc(),
            )
            .all()
        )

        items = []

        for loc in locations:

            items.append({
                "id": loc.id,
                "code": loc.code or "",
                "description": loc.description or "",
            })

        return {
            "items": items
        }

    except Exception as exc:

        print(f"[WAREHOUSE LOCATIONS ERROR] {exc}")

        return {
            "items": []
        }, 500

@purchases_bp.route("/inventory-entries/<int:entry_id>")
@login_required
def entry_detail(entry_id: int):
    inventory_entry = get_inventory_entry_or_404(entry_id)
    return render_template(
        "purchases/inventory_entries/detail.html",
        inventory_entry=inventory_entry,
    )

@purchases_bp.route("/quotations/line/<int:line_id>")
@login_required
def quotation_line_view(line_id: int):
    request_line = PurchaseRequestLine.query.get_or_404(line_id)

    try:
        comparison = get_comparison_for_purchase_request_line(
            purchase_request_line_id=line_id
        )
    except QuotationServiceError as exc:
        flash(str(exc), "danger")

        if request_line.purchase_request_id:
            return redirect(
                url_for(
                    "purchases.quotation_request_lines",
                    request_id=request_line.purchase_request_id,
                )
            )

        return redirect(url_for("purchases.list_quotations"))

    article_id = request_line.article_id

    suppliers = []

    if article_id:
        supplier_ids = [
            row.supplier_id
            for row in (
                ArticleSupplier.query
                .filter(
                    ArticleSupplier.article_id == article_id,
                    ArticleSupplier.is_active.is_(True),
                )
                .all()
            )
        ]

        if supplier_ids:
            suppliers = (
                Supplier.query
                .filter(
                    Supplier.id.in_(supplier_ids),
                    Supplier.is_active.is_(True),
                )
                .order_by(
                    Supplier.commercial_name.asc(),
                    Supplier.legal_name.asc(),
                )
                .all()
            )

    return render_template(
        "purchases/quotations/line.html",
        line_id=line_id,
        request_line=request_line,
        comparison=comparison,
        suppliers=suppliers,
    )

@purchases_bp.route("/quotations/line/<int:line_id>/quote", methods=["POST"])
@login_required
def create_quote_for_line(line_id: int):
    supplier_id = _to_int(request.form.get("supplier_id"))
    new_supplier_name = request.form.get("new_supplier_name")

    if supplier_id == -1:
        supplier_id = None
    else:
        new_supplier_name = None

    unit_price = request.form.get("unit_price")
    use_last_price = request.form.get("use_last_price") == "1"

    discount_pct = request.form.get("discount_pct") or "0"
    tax_pct = request.form.get("tax_pct") or "13"
    tax_included = request.form.get("tax_included") == "1"

    payment_type = request.form.get("payment_type") or None
    payment_term = _to_int(request.form.get("payment_term_months"))
    origin_type = request.form.get("origin_type") or None

    lead_time_days = _to_int(request.form.get("lead_time_days"))
    brand_model = request.form.get("brand_model")
    notes = request.form.get("notes")
    status = request.form.get("status") or "COTIZADA"

    try:
        create_single_line_quotation(
            purchase_request_line_id=line_id,
            supplier_id=supplier_id,
            new_supplier_name=new_supplier_name,
            created_by_user_id=current_user.id,
            unit_price=unit_price,
            use_last_price=use_last_price,
            discount_pct=discount_pct,
            tax_pct=tax_pct,
            tax_included=tax_included,
            payment_type=payment_type,
            payment_term_months=payment_term,
            origin_type=origin_type,
            lead_time_days=lead_time_days,
            brand_model=brand_model,
            notes=notes,
            status=status,
        )
    except QuotationServiceError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("purchases.quotation_line_view", line_id=line_id))

    flash("Cotización guardada correctamente.", "success")
    request_line = PurchaseRequestLine.query.get(line_id)

    return redirect(
        url_for(
            "purchases.quotation_request_lines",
            request_id=request_line.purchase_request_id,
        )
    )

@purchases_bp.route("/quotations/last-price")
@login_required
def get_last_price():
    supplier_id = _to_int(request.args.get("supplier_id"))
    article_id = _to_int(request.args.get("article_id"))
    pending_article_id = _to_int(request.args.get("pending_article_id"))

    try:
        last = get_last_price_for_supplier(
            supplier_id=supplier_id,
            article_id=article_id,
            pending_article_id=pending_article_id,
        )
    except QuotationServiceError as exc:
        return {"error": str(exc)}, 400

    if not last:
        return {"price": None}

    return {
        "price": str(last.unit_price),
        "date": str(last.quote_date),
    }

@purchases_bp.route("/quotations/line/<int:line_id>/export-excel")
@login_required
def export_quotation_comparison_excel(line_id: int):
    try:
        comparison = get_comparison_for_purchase_request_line(
            purchase_request_line_id=line_id
        )
    except QuotationServiceError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("purchases.create_quotation"))

    request_line = PurchaseRequestLine.query.get_or_404(line_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparativo"

    title = "Comparativo de Cotizaciones"
    ws.merge_cells("A1:L1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A3"] = "Artículo:"
    ws["B3"] = request_line.item_name or "Sin artículo"
    ws["A4"] = "Código:"
    ws["B4"] = request_line.item_code or "-"
    ws["A5"] = "Cantidad:"
    ws["B5"] = request_line.quantity_requested

    headers = [
        "#",
        "Proveedor",
        "Monto ingresado",
        "Tipo IVA",
        "Subtotal",
        "Descuento %",
        "Monto descuento",
        "IVA %",
        "Monto IVA",
        "Total",
        "Fecha",
        "Pago",
    ]

    start_row = 7

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.alignment = Alignment(horizontal="center")

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for index, item in enumerate(comparison, start=1):
        row = start_row + index

        values = [
            "MEJOR" if item.get("is_best_price") else item.get("rank"),
            item.get("supplier_name"),
            float(item.get("last_price") or 0),
            "Con IVA" if item.get("tax_included") else "Sin IVA",
            float(item.get("subtotal") or 0),
            float(item.get("discount_pct") or 0),
            float(item.get("discount_amount") or 0),
            float(item.get("tax_pct") or 0),
            float(item.get("tax_amount") or 0),
            float(item.get("total_amount") or 0),
            item.get("last_quote_date").strftime("%d/%m/%Y") if item.get("last_quote_date") else "-",
            (
                f"{item.get('payment_type')} / {item.get('payment_term_months')} meses"
                if item.get("payment_type") and item.get("payment_term_months")
                else item.get("payment_type") or "-"
            ),
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center")

            if col in [3, 5, 7, 9, 10]:
                cell.number_format = '#,##0.00'

            if col in [6, 8]:
                cell.number_format = '0.00'

            if item.get("is_best_price"):
                cell.fill = PatternFill("solid", fgColor="DCFCE7")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    ws.freeze_panes = "A8"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"comparativo_cotizacion_linea_{line_id}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@purchases_bp.route("/quotations/line/<int:line_id>/print")
@login_required
def print_quotation_comparison(line_id: int):
    try:
        comparison = get_comparison_for_purchase_request_line(
            purchase_request_line_id=line_id
        )
    except QuotationServiceError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("purchases.create_quotation"))

    request_line = PurchaseRequestLine.query.get_or_404(line_id)

    return render_template(
        "purchases/quotations/print_comparison.html",
        request_line=request_line,
        comparison=comparison,
        generated_at=datetime.now(CR_TZ),
    )

@purchases_bp.route("/orders/<int:order_id>/upload-approved-pdf", methods=["POST"])
@login_required
def upload_approved_pdf(order_id):
    file = request.files.get("file")

    if not file or not file.filename:
        flash("Debe subir un archivo PDF.", "danger")
        return redirect(url_for("purchases.order_detail", order_id=order_id))

    if not file.filename.lower().endswith(".pdf"):
        flash("El archivo debe ser PDF.", "danger")
        return redirect(url_for("purchases.order_detail", order_id=order_id))

    order = PurchaseOrder.query.get_or_404(order_id)

    file_data = file.read()

    if not file_data:
        flash("El archivo PDF está vacío.", "danger")
        return redirect(url_for("purchases.order_detail", order_id=order_id))

    approval = PurchaseOrderApproval(
        purchase_order_id=order_id,
        approved_by_user_id=current_user.id,
        status="APROBADA",
        reason="OC firmada y adjuntada en PDF.",
        approved_pdf_data=file_data,
        approved_pdf_mime_type=file.mimetype or "application/pdf",
        approved_pdf_original_name=file.filename,
        approved_pdf_uploaded_at=datetime.now(UTC),
    )

    order.approval_status = "APROBADA"
    order.approved_at = datetime.now(UTC)

    db.session.add(approval)
    db.session.commit()

    flash("Orden aprobada correctamente con PDF firmado.", "success")
    return redirect(url_for("purchases.order_detail", order_id=order_id))


@purchases_bp.route("/orders/<int:order_id>/approved-pdf")
@login_required
def view_approved_pdf(order_id):
    approval = (
        PurchaseOrderApproval.query
        .filter(
            PurchaseOrderApproval.purchase_order_id == order_id,
            PurchaseOrderApproval.status == "APROBADA",
            PurchaseOrderApproval.approved_pdf_data.isnot(None),
        )
        .order_by(PurchaseOrderApproval.created_at.desc())
        .first()
    )

    if not approval:
        flash("No hay PDF aprobado para esta orden.", "warning")
        return redirect(url_for("purchases.order_detail", order_id=order_id))

    filename = approval.approved_pdf_original_name or f"OC_{order_id}.pdf"

    return Response(
        approval.approved_pdf_data,
        mimetype=approval.approved_pdf_mime_type or "application/pdf",
        headers={
            "Content-Disposition": f'inline; filename="{filename}"'
        },
    )

@purchases_bp.route(
    "/orders/lines/<int:line_id>/adjust",
    methods=["POST"],
)
@login_required
def adjust_purchase_order_line(line_id: int):

    quantity = request.form.get("quantity_ordered")
    unit_cost = request.form.get("unit_cost")

    try:

        adjusted_line = adjust_approved_purchase_order_line(
            purchase_order_line_id=line_id,
            new_quantity=quantity,
            new_unit_cost=unit_cost,
        )

    except PurchaseOrderServiceError as exc:

        flash(str(exc), "danger")

        return redirect(
            url_for(
                "purchases.order_detail",
                order_id=PurchaseOrderLine.query.get_or_404(line_id).purchase_order_id,
            )
        )

    flash(
        "Línea de orden ajustada correctamente.",
        "success",
    )

    return redirect(
        url_for(
            "purchases.order_detail",
            order_id=adjusted_line.purchase_order_id,
        )
    )