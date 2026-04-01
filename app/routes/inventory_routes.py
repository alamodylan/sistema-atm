from __future__ import annotations

from io import BytesIO

from flask import (
    Blueprint,
    flash,
    redirect,
    render_template,
    send_file,
    session,
    url_for,
)
from flask_login import login_required
from openpyxl import Workbook
from openpyxl.styles import Font

from app.models.warehouse import Warehouse
from app.services.inventory_service import (
    InventoryServiceError,
    get_inventory_by_warehouse,
    get_structures_by_site_and_type,
)

inventory_bp = Blueprint("inventory", __name__)

WAREHOUSE_TYPE_LABELS = {
    "BODEGA": "Bodegas",
    "MINIBODEGA": "Mini bodegas",
    "CAJA_HERRAMIENTAS": "Cajas de herramientas",
}


def _get_active_site_id() -> int:
    active_site_id = session.get("active_site_id")
    if not active_site_id:
        raise InventoryServiceError(
            "Debe seleccionar un predio antes de ingresar al inventario."
        )
    return int(active_site_id)


def _get_valid_warehouse_type(warehouse_type: str) -> str:
    warehouse_type = (warehouse_type or "").strip().upper()
    if warehouse_type not in WAREHOUSE_TYPE_LABELS:
        raise InventoryServiceError("El tipo de inventario solicitado no es válido.")
    return warehouse_type


# =========================================================
# INVENTARIO - PANTALLA PRINCIPAL
# =========================================================
@inventory_bp.route("/", methods=["GET"])
@login_required
def inventory_home():
    try:
        active_site_id = _get_active_site_id()

        bodegas = get_structures_by_site_and_type(active_site_id, "BODEGA")
        minibodegas = get_structures_by_site_and_type(active_site_id, "MINIBODEGA")
        cajas = get_structures_by_site_and_type(active_site_id, "CAJA_HERRAMIENTAS")

        return render_template(
            "inventory/index.html",
            title="Inventario",
            subtitle="Seleccione el tipo de estructura para consultar el inventario del predio activo.",
            active_site_id=active_site_id,
            active_type="BODEGA",
            structures_by_type={
                "BODEGA": bodegas,
                "MINIBODEGA": minibodegas,
                "CAJA_HERRAMIENTAS": cajas,
            },
            type_labels=WAREHOUSE_TYPE_LABELS,
        )

    except InventoryServiceError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("dashboard.home"))

    except Exception:
        flash("Error al cargar la pantalla principal de inventario.", "danger")
        return redirect(url_for("dashboard.home"))


# =========================================================
# INVENTARIO - ESTRUCTURAS POR TIPO
# =========================================================
@inventory_bp.route("/type/<string:warehouse_type>", methods=["GET"])
@login_required
def inventory_type(warehouse_type: str):
    try:
        active_site_id = _get_active_site_id()
        warehouse_type = _get_valid_warehouse_type(warehouse_type)

        structures = get_structures_by_site_and_type(active_site_id, warehouse_type)

        return render_template(
            "inventory/index.html",
            title="Inventario",
            subtitle="Seleccione la estructura que desea consultar dentro del predio activo.",
            active_site_id=active_site_id,
            active_type=warehouse_type,
            structures_by_type={
                "BODEGA": get_structures_by_site_and_type(active_site_id, "BODEGA"),
                "MINIBODEGA": get_structures_by_site_and_type(active_site_id, "MINIBODEGA"),
                "CAJA_HERRAMIENTAS": get_structures_by_site_and_type(active_site_id, "CAJA_HERRAMIENTAS"),
            },
            selected_structures=structures,
            type_labels=WAREHOUSE_TYPE_LABELS,
        )

    except InventoryServiceError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("inventory.inventory_home"))

    except Exception:
        flash("Error al cargar las estructuras del inventario.", "danger")
        return redirect(url_for("inventory.inventory_home"))


# =========================================================
# INVENTARIO - DETALLE DE UNA ESTRUCTURA
# =========================================================
@inventory_bp.route("/warehouse/<int:warehouse_id>", methods=["GET"])
@login_required
def warehouse_inventory_detail(warehouse_id: int):
    try:
        active_site_id = _get_active_site_id()

        warehouse = Warehouse.query.get(warehouse_id)
        if not warehouse:
            raise InventoryServiceError("La estructura solicitada no existe.")

        if int(warehouse.site_id) != active_site_id:
            raise InventoryServiceError("La estructura solicitada no pertenece al predio activo.")

        items = get_inventory_by_warehouse(warehouse_id)

        return render_template(
            "inventory/detail.html",
            title="Inventario de estructura",
            subtitle="Consulte el inventario real de la estructura seleccionada.",
            warehouse=warehouse,
            items=items,
            active_site_id=active_site_id,
            warehouse_type_label=WAREHOUSE_TYPE_LABELS.get(
                warehouse.warehouse_type, warehouse.warehouse_type
            ),
        )

    except InventoryServiceError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("inventory.inventory_home"))

    except Exception:
        flash("Error al cargar el detalle del inventario.", "danger")
        return redirect(url_for("inventory.inventory_home"))


# =========================================================
# INVENTARIO - EXPORTAR A EXCEL
# =========================================================
@inventory_bp.route("/warehouse/<int:warehouse_id>/export", methods=["GET"])
@login_required
def export_warehouse_inventory(warehouse_id: int):
    try:
        active_site_id = _get_active_site_id()

        warehouse = Warehouse.query.get(warehouse_id)
        if not warehouse:
            raise InventoryServiceError("La estructura solicitada no existe.")

        if int(warehouse.site_id) != active_site_id:
            raise InventoryServiceError("La estructura solicitada no pertenece al predio activo.")

        items = get_inventory_by_warehouse(warehouse_id)

        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"

        headers = [
            "Código",
            "Artículo",
            "Existencia",
            "Costo último",
            "Costo promedio",
        ]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        for item in items:
            ws.append(
                [
                    item["code"],
                    item["name"],
                    float(item["quantity_on_hand"]),
                    float(item["last_unit_cost"]),
                    float(item["avg_unit_cost"]),
                ]
            )

        ws["G1"] = "Estructura"
        ws["G2"] = warehouse.name
        ws["H1"] = "Código estructura"
        ws["H2"] = warehouse.code
        ws["I1"] = "Tipo"
        ws["I2"] = warehouse.warehouse_type

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"inventario_{warehouse.code}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except InventoryServiceError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("inventory.inventory_home"))

    except Exception:
        flash("Error al exportar el inventario a Excel.", "danger")
        return redirect(url_for("inventory.inventory_home"))