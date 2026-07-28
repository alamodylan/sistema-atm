from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


YELLOW_FILL = PatternFill("solid", fgColor="FFF200")
BEST_PRICE_FILL = PatternFill("solid", fgColor="92D050")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
THIN_BLACK = Side(style="thin", color="000000")
TABLE_BORDER = Border(
    left=THIN_BLACK,
    right=THIN_BLACK,
    top=THIN_BLACK,
    bottom=THIN_BLACK,
)


def _excel_number(value: Any) -> float | None:
    """Convierte Decimal u otro número a un valor numérico para Excel."""
    if value is None:
        return None

    try:
        return float(Decimal(str(value)))
    except Exception:
        return None


def _clean_filename(value: str) -> str:
    """Limpia caracteres problemáticos para nombres de archivos."""
    cleaned = str(value or "").strip()

    for character in ('/', '\\', ':', '*', '?', '"', '<', '>', '|'):
        cleaned = cleaned.replace(character, "-")

    return cleaned or "comparativo"


def build_category_comparison_workbook(
    *,
    comparison_matrix: dict,
) -> Workbook:
    """
    Genera un comparativo horizontal por categoría.

    Estructura:
    Cód. | Descripción | CANT. | Proveedor 1 | Proveedor 2 | ...

    Cada celda de proveedor muestra el total unitario cotizado.
    La mejor oferta se resalta en verde según la lógica ya calculada
    por get_category_comparison_matrix().
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Comparativo"
    sheet.sheet_view.showGridLines = False

    suppliers = comparison_matrix.get("suppliers") or []
    lines = comparison_matrix.get("lines") or []

    headers = ["Cód.", "Descripción", "CANT."]
    headers.extend(
        supplier.get("supplier_name") or f"Proveedor {supplier.get('supplier_id')}"
        for supplier in suppliers
    )

    # =====================================================
    # ENCABEZADOS
    # =====================================================
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.font = Font(bold=True, color="000000")
        cell.fill = YELLOW_FILL
        cell.border = TABLE_BORDER
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    sheet.row_dimensions[1].height = 30

    supplier_column_by_id = {
        supplier.get("supplier_id"): index
        for index, supplier in enumerate(suppliers, start=4)
    }

    # =====================================================
    # FILAS DEL COMPARATIVO
    # =====================================================
    for row_index, line in enumerate(lines, start=2):
        code_cell = sheet.cell(
            row=row_index,
            column=1,
            value=line.get("item_code") or "-",
        )
        description_cell = sheet.cell(
            row=row_index,
            column=2,
            value=line.get("item_name") or "Sin artículo",
        )
        quantity_cell = sheet.cell(
            row=row_index,
            column=3,
            value=_excel_number(line.get("quantity_requested")) or 0,
        )

        code_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        description_cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )
        quantity_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        quantity_cell.number_format = '#,##0.####'

        for cell in (code_cell, description_cell, quantity_cell):
            cell.border = TABLE_BORDER
            cell.fill = WHITE_FILL

        cells_by_supplier_id = {
            cell.get("supplier_id"): cell
            for cell in (line.get("cells") or [])
        }

        for supplier in suppliers:
            supplier_id = supplier.get("supplier_id")
            column_index = supplier_column_by_id[supplier_id]
            matrix_cell = cells_by_supplier_id.get(supplier_id)

            excel_cell = sheet.cell(row=row_index, column=column_index)
            excel_cell.border = TABLE_BORDER
            excel_cell.alignment = Alignment(
                horizontal="right",
                vertical="center",
            )

            if not matrix_cell or not matrix_cell.get("has_quotation"):
                excel_cell.value = "-"
                excel_cell.fill = WHITE_FILL
                excel_cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )
                continue

            quotation = matrix_cell.get("quotation") or {}
            value = _excel_number(quotation.get("unit_total"))

            if value is None:
                excel_cell.value = "-"
                excel_cell.fill = WHITE_FILL
                excel_cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )
                continue

            excel_cell.value = value
            excel_cell.number_format = '#,##0.00'
            excel_cell.fill = (
                BEST_PRICE_FILL
                if quotation.get("is_best_price")
                else WHITE_FILL
            )

            currency_code = quotation.get("currency_code") or ""
            quote_date = quotation.get("quote_date")
            comment_parts = []

            if currency_code:
                comment_parts.append(f"Moneda: {currency_code}")

            if quote_date:
                try:
                    comment_parts.append(
                        f"Fecha: {quote_date.strftime('%d/%m/%Y')}"
                    )
                except Exception:
                    comment_parts.append(f"Fecha: {quote_date}")

            if quotation.get("brand_model"):
                comment_parts.append(
                    f"Marca / modelo: {quotation.get('brand_model')}"
                )

            if comment_parts:
                excel_cell.comment = Comment(
                    "\n".join(comment_parts),
                    "Sistema de Compras",
                )

        sheet.row_dimensions[row_index].height = 23

    # =====================================================
    # PRESENTACIÓN Y CONFIGURACIÓN DE IMPRESIÓN
    # =====================================================
    sheet.column_dimensions["A"].width = 12
    sheet.column_dimensions["B"].width = 38
    sheet.column_dimensions["C"].width = 11

    for column_index in range(4, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(column_index)].width = 18

    last_column_letter = get_column_letter(max(len(headers), 3))
    last_row = max(len(lines) + 1, 1)

    sheet.auto_filter.ref = f"A1:{last_column_letter}{last_row}"
    sheet.freeze_panes = "D2"
    sheet.print_title_rows = "1:1"
    sheet.print_area = f"A1:{last_column_letter}{last_row}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.40
    sheet.page_margins.bottom = 0.40
    sheet.sheet_view.zoomScale = 85

    return workbook


def build_category_comparison_excel(
    *,
    comparison_matrix: dict,
) -> BytesIO:
    """Construye el archivo Excel y lo devuelve listo para send_file()."""
    workbook = build_category_comparison_workbook(
        comparison_matrix=comparison_matrix,
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def category_comparison_filename(
    *,
    purchase_request_number: str | None,
    category_label: str | None,
) -> str:
    request_part = _clean_filename(
        purchase_request_number or "solicitud"
    )
    category_part = _clean_filename(
        category_label or "sin-categoria"
    )

    return f"comparativo_{request_part}_{category_part}.xlsx"